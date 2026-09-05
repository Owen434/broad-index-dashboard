# -*- coding: utf-8 -*-
"""
A股个股分析总控脚本（形态 + 八大指标矩阵 + 波段看板 + 评分矩阵 四合一）
==========================================================================
把原来分散在四个脚本里的能力合并成一条流水线，分析对象由「基金」改为「个股」：

    ztjj_board_stocks.csv  ──勾选板块/概念──▶  取数(Price_movement_patterns)
            │
            ├─▶ ① K线形态与缺口 HTML + 板块形态统计汇总.xlsx   （原 Price_movement_patterns.py）
            ├─▶ ② 近60日八大指标交互式矩阵                      （原 17_Fund_6Indicators_Matrix.py）
            ├─▶ ③ ZigZag 波浪 + 波段信号量化看板                （原 18_zigzag_signal_analyzer.py）
            └─▶ ④ 过热评分矩阵 + T+1 涨跌情景模拟               （原 19_Fund_ScoreMatrix.py）

【设计原则】
1. 指标算法绝不重写：八大指标 / 过热评分 / 状态分档 / ZigZag / 图表构建
   仍然 100% 调用 18_zigzag_signal_analyzer.py 里的函数；
   取数 / 形态识别 / 缺口 / 涨跌停 / 量价 仍然 100% 调用 Price_movement_patterns.py 里的函数。
   本脚本只负责「串流程 + 出页面」，阈值改一处全局生效的架构不变。
2. 基金版按 持有/板块/自选 分类 → 个股版改按 板块类别（行业 / 概念）分类，
   并额外提供「具体板块」下拉筛选（一只票横跨多个板块时，几个分类都能筛到它）。
3. 基金版要先做净值复权(build_adjusted_nav)；个股走 akshare 前复权(qfq)接口，
   本身已是连续可比序列，故跳过复权环节，其余口径完全一致。

【运行前提】
    price_movement_patterns.py      取数 + 形态内核（不在本仓库里，需自行提供）
    ../funds/zigzag_signal_analyzer.py  指标 + 评分 + 绘图内核（本仓库自带）
    ztjj_board_stocks.csv           板块成分股，仅在 BROAD_INDEX_ONLY=False 时才需要
    target_stocks.csv               可选，仅用于并入国外标的

【开源版说明】默认 BROAD_INDEX_ONLY=True：只跑十来个宽基指数，不需要上面两个 CSV，
输出三份 HTML 只含宽基数据，不含任何个股。想恢复"按板块勾选个股"的完整功能，
把 BROAD_INDEX_ONLY 改成 False，并自备 price_movement_patterns.py 与两份 CSV。
"""

import os
import re
import sys
import json
import time
import importlib.util
import warnings
import multiprocessing as mp
from concurrent.futures import ThreadPoolExecutor, ProcessPoolExecutor, as_completed

import pandas as pd
import numpy as np

warnings.filterwarnings('ignore')

# ==========================================================================
# 0. 配置区
# ==========================================================================
# ---- 内核脚本 ----
PM_FILE_CANDIDATES = ['price_movement_patterns.py', 'price_Movement_Patterns.py']
CORE_FILE_CANDIDATES = ['zigzag_signal_analyzer.py']
# 两个内核不在同一目录时，把目录写在这里（本脚本目录与当前工作目录会自动搜索）。
# 仓库里 zigzag_signal_analyzer.py 放在 ../funds/ 下，所以默认写这一条相对路径；
# price_movement_patterns.py（K线形态/取数内核）不在本仓库里，需要你自己补上并把目录加进来。
EXTRA_CORE_DIRS = [
    os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "funds"),
]

# ---- 输入 ----
BOARD_CSV = "ztjj_board_stocks.csv"     # 板块成分股：板块名称/板块类别/股票代码/股票名称/基金持有比例(%)
LEGACY_CSV = "target_stocks.csv"        # 老股票池：仅用于取「国外」标的

# ---- 输出 ----
CHART_DIR = "stock_charts_output"       # K线形态图 + 板块形态统计汇总.xlsx
DASHBOARD_DIR = "stock_charts_output"   # 三张看板 HTML 的落地目录，想放当前目录改成 "."
HTML_MATRIX = "stock_8indicators_matrix.html"
HTML_ZIGZAG = "stock_zigzag_signal_analyzer.html"
HTML_SCORE = "stock_scorematrix.html"

# 开源版：只跑宽基指数（约十来个），不勾选个股板块，也不需要 ztjj_board_stocks.csv。
# 想恢复"按板块挑选个股"的完整功能（需要你自己的板块成分股 CSV），把这里改成 False。
BROAD_INDEX_ONLY = True

# ---- 选股默认值 ----
USE_SELECTION_UI = True                 # 是否弹出勾选对话框（False = 静默用下面的默认值，适合定时任务）
DEFAULT_BOARD_CATEGORY = "概念"         # 对话框默认类别：概念 / 行业 / 全部
DEFAULT_TOP_N_PER_BOARD = 10            # ★ 默认每个板块取「基金持有比例(%)」前 10 只（0 = 全部）
TOP_N_PRESETS = [5, 10, 20, 30, 50, 0]  # 对话框上的快捷按钮（0 显示为「全部」）
DEFAULT_BOARDS = None                   # None = 沿用 Price_movement_patterns.py 里的 DEFAULT_BOARDS

INCLUDE_OVERSEAS = False                # 是否并入 target_stocks.csv 的国外标的
INCLUDE_LEGACY_DOMESTIC = False         # 是否并入 target_stocks.csv 的国内标的
INCLUDE_BROAD_INDICES = True            # 是否并入宽基指数（Excel 的「宽基」表靠它）

# ---- 输出模块开关（对话框里可勾）----
OUT_PATTERN_CHARTS = True               # ① 每只票一张 K线形态与缺口 HTML
OUT_PATTERN_EXCEL = True                # ① 板块形态统计汇总.xlsx
OUT_MATRIX = True                    # ② 八大指标矩阵
OUT_ZIGZAG = True                    # ③ 波段信号量化看板
OUT_SCORE = True                     # ④ 评分矩阵 + T+1 情景模拟

# ---- 并发 ----
FETCH_WORKERS = 8                       # 抓数据并发线程数（太大易被接口限流）
FETCH_RETRY = 2
RENDER_WORKERS = max(1, (os.cpu_count() or 4) - 1)
USE_PROCESS_POOL = True                 # 出问题时置 False 退回单进程

# ---- 计算口径 ----
# 指标 / 评分只取最近 N 根K线参与计算（0 = 全部历史）。
# A股老票动辄 5000+ 根，全量跑 13 次 T+1 情景重算会非常慢；
# 1200 根 ≈ 5 年，足够覆盖 18 号看板「近1年分位」与「全部分位」的常用视角。
# 想要严格的全历史分位数，把它设成 0。
ANALYSIS_MAX_BARS = 1200
SHOW_HIST_DAYS = 60                     # 矩阵页历史列最多显示多少个交易日
SIM_RETURNS = [6.0, 5.0, 4.0, 3.0, 2.0, 1.0, 0.0, -1.0, -2.0, -3.0, -4.0, -5.0, -6.0]
CELL_SHOW_INDICATOR_TIP = True          # 评分矩阵单元格是否附带八大指标 hover 提示

# ZigZag 看板会把每只票的两张 Plotly 图内嵌进同一个 HTML，标的一多页面会非常臃肿。
# 这里按「基金持有比例」从高到低截断（0 = 不限制）。矩阵页与 Excel 不受该上限影响。
ZIGZAG_MAX_TARGETS = 100
# 宽基指数是否免于上限：True = 十来个指数照单全收、不挤占个股名额
ZIGZAG_INDEX_ALWAYS = True

# ==========================================================================
# 1. 内核脚本加载（父/子进程各自懒加载，保证进程池下也能用）
# ==========================================================================
_PM = None
_CORE = None


def _search_dirs():
    dirs = []
    try:
        dirs.append(os.path.dirname(os.path.abspath(__file__)))
    except NameError:
        pass
    dirs.append(os.getcwd())
    dirs.extend(EXTRA_CORE_DIRS)
    seen, out = set(), []
    for d in dirs:
        if d and d not in seen and os.path.isdir(d):
            seen.add(d)
            out.append(d)
    return out


def _load_module(alias, candidates, human_name):
    """按候选文件名在若干目录里找脚本并加载。两个内核的主程序都写在
    __name__ == '__main__' 保护里，import 只会拿到函数定义，不会触发它们生成文件。"""
    if alias in sys.modules:
        return sys.modules[alias]
    for d in _search_dirs():
        for fname in candidates:
            path = os.path.join(d, fname)
            if not os.path.exists(path):
                continue
            spec = importlib.util.spec_from_file_location(alias, path)
            module = importlib.util.module_from_spec(spec)
            sys.modules[alias] = module
            spec.loader.exec_module(module)
            return module
    raise SystemExit(
        f"错误：未找到{human_name}。请把下列任一文件放到本脚本同目录，"
        f"或写进 EXTRA_CORE_DIRS：\n  " + '\n  '.join(candidates)
    )


def pm():
    """取数 + 形态内核（Price_movement_patterns.py）"""
    global _PM
    if _PM is None:
        _PM = _load_module('price_movement_core', PM_FILE_CANDIDATES, '取数/形态内核')
    return _PM


def core():
    """指标 + 评分 + 绘图内核（18_zigzag_signal_analyzer.py）"""
    global _CORE
    if _CORE is None:
        _CORE = _load_module('zigzag_indicator_core', CORE_FILE_CANDIDATES, '指标内核')
    return _CORE


# ==========================================================================
# 2. 小工具
# ==========================================================================
def short_code(code):
    """sh600519 -> 600519；美股 / 指数代码原样返回，用于页面展示"""
    c = str(code)
    m = re.match(r'^(sh|sz|bj)(\d{6})$', c)
    return m.group(2) if m else c


def safe_key(code):
    """HTML id / JS 键：只保留字母数字与下划线（.IXIC 这类代码不能直接进 id）"""
    return re.sub(r'[^0-9A-Za-z_]', '_', str(code))


def _fmt(value, spec, suffix=''):
    """NaN 统一显示成 --，避免表格里出现 nan。"""
    if value is None or (isinstance(value, float) and np.isnan(value)):
        return '--'
    return format(value, spec) + suffix


def _hex_to_rgba(hex_color, alpha):
    """#RRGGBB -> rgba(r,g,b,alpha)，用于评分徽章的淡色底。"""
    h = str(hex_color).lstrip('#')
    if len(h) == 3:
        h = ''.join(c * 2 for c in h)
    if len(h) != 6:
        return f'rgba(136,136,136,{alpha})'
    r, g, b = int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
    return f'rgba({r},{g},{b},{alpha})'


def _esc(text):
    """写进 HTML 属性前的最小转义。"""
    return (str(text).replace('&', '&amp;').replace('<', '&lt;')
            .replace('>', '&gt;').replace('"', '&quot;'))


def _indicator_tip(row):
    """把八大指标压成一行 hover 文本。"""
    if not CELL_SHOW_INDICATOR_TIP or row is None:
        return ''
    bw = row.get('boll_bw_pct', np.nan)
    bw = np.nan if bw is None else float(bw)
    parts = [
        f"MACD {_fmt(float(row.get('macd_hist', np.nan)), '+.3f')}",
        f"RSI {_fmt(float(row.get('rsi', np.nan)), '.1f')}",
        f"BIAS {_fmt(float(row.get('bias', np.nan)), '+.2f', '%')}",
        f"KDJ {_fmt(float(row.get('kdj', np.nan)), '.1f')}",
        f"CCI {_fmt(float(row.get('cci', np.nan)), '.1f')}",
        f"SAR {row.get('sar_stat', '--')}",
        f"BOLL {_fmt(float(row.get('boll_pctb', np.nan)), '.3f')}",
        f"BW% {_fmt(np.nan if (bw is None or np.isnan(bw)) else bw * 100, '.0f', '%')}",
    ]
    return ' | '.join(parts).replace('"', "'")


def _ind_row_from_series(s):
    """把 compute_eight_indicators 的最后一行转成本脚本内部列名的 dict。"""
    sar_val = s.get('sar_bullish', np.nan)
    try:
        sar_txt = '看涨' if bool(sar_val) else '看跌'
    except Exception:
        sar_txt = '--'
    return {
        'macd_hist': s.get('macd', np.nan),
        'rsi': s.get('rsi', np.nan),
        'bias': s.get('bias', np.nan),
        'kdj': s.get('kdj', np.nan),
        'cci': s.get('cci', np.nan),
        'sar_stat': sar_txt,
        'boll_pctb': s.get('boll_pctb', np.nan),
        'boll_bw_pct': s.get('bw_pct', np.nan),
    }


def simulate_next_day(close, pct):
    """在收盘价序列尾部追加一个「下一交易日」的模拟价：
           新价 = 最新价 × (1 + pct/100)
    然后完整重算八大指标与过热评分，返回 (指标dict, 评分float或None)。

    评分链路 compute_heat_score_series -> calc_score_series 与实盘完全一致，
    所以模拟值和真实次日收盘后跑出来的结果是同一套口径。
    """
    c = core()
    s = pd.Series(close).dropna()
    if len(s) < c.MIN_ROWS:
        return None, None
    new_val = float(s.iloc[-1]) * (1.0 + pct / 100.0)

    last_date = s.index[-1]
    try:
        new_date = last_date + pd.tseries.offsets.BDay(1)
    except Exception:
        new_date = last_date + 1

    sim = pd.concat([s, pd.Series([new_val], index=[new_date])])
    sim = sim[~sim.index.duplicated(keep='last')].sort_index()

    ind = c.compute_eight_indicators(sim)
    net_s, int_s, _, _ = c.compute_heat_score_series(ind)
    score_s = c.calc_score_series(net_s, int_s)

    last_score = score_s.iloc[-1]
    return _ind_row_from_series(ind.iloc[-1]), (None if pd.isna(last_score) else float(last_score))


# ---- 状态定义（与 18 号内核的 get_status_action 一一对应）----
STATUS_ORDER = ['🔴 极端风险', '🟠 高风险', '🟡 偏热', '🟢 正常', '🟦 偏冷', '🔵 冰点', '⚪ 无数据']
STATUS_COLOR_MAP = {
    '🔴 极端风险': '#EF4444',
    '🟠 高风险': '#F97316',
    '🟡 偏热': '#EAB308',
    '🟢 正常': '#22C55E',
    '🟦 偏冷': '#60A5FA',
    '🔵 冰点': '#3B82F6',
    '⚪ 无数据': '#888888',
}
STATUS_SLUG_MAP = {
    '🔴 极端风险': 'extreme',
    '🟠 高风险': 'high',
    '🟡 偏热': 'hot',
    '🟢 正常': 'normal',
    '🟦 偏冷': 'cool',
    '🔵 冰点': 'cold',
    '⚪ 无数据': 'nodata',
}


# ==========================================================================
# 3. 板块勾选对话框（在原对话框基础上，新增 Top N 快捷按钮 + 输出模块按钮）
# ==========================================================================
def _default_boards():
    return DEFAULT_BOARDS if DEFAULT_BOARDS is not None else pm().DEFAULT_BOARDS


def _base_choice():
    """所有选项的默认值，对话框与命令行都以它为起点。"""
    return dict(
        boards=[], top_n=DEFAULT_TOP_N_PER_BOARD, category=DEFAULT_BOARD_CATEGORY,
        overseas=INCLUDE_OVERSEAS, legacy=INCLUDE_LEGACY_DOMESTIC, broad=INCLUDE_BROAD_INDICES,
        charts=OUT_PATTERN_CHARTS, excel=OUT_PATTERN_EXCEL,
        m17=OUT_MATRIX, m18=OUT_ZIGZAG, m19=OUT_SCORE,
    )


def select_options_ui(pool, catalog):
    """弹出勾选对话框；无图形界面则返回 None 交给命令行版本"""
    try:
        import tkinter as tk
        from tkinter import ttk, messagebox
    except Exception:
        return None
    try:
        root = tk.Tk()
    except Exception:
        return None   # 服务器 / 无显示环境

    P = pm()
    result = {}
    root.title("选择要分析的板块与输出内容")
    root.geometry("1180x780")
    root.minsize(860, 620)

    cat_var = tk.StringVar(value=DEFAULT_BOARD_CATEGORY if DEFAULT_BOARD_CATEGORY in ('概念', '行业', '全部') else '概念')
    kw_var = tk.StringVar()
    topn_var = tk.StringVar(value=str(DEFAULT_TOP_N_PER_BOARD))
    ovs_var = tk.BooleanVar(value=INCLUDE_OVERSEAS)
    legacy_var = tk.BooleanVar(value=INCLUDE_LEGACY_DOMESTIC)
    broad_var = tk.BooleanVar(value=INCLUDE_BROAD_INDICES)
    charts_var = tk.BooleanVar(value=OUT_PATTERN_CHARTS)
    excel_var = tk.BooleanVar(value=OUT_PATTERN_EXCEL)
    m17_var = tk.BooleanVar(value=OUT_MATRIX)
    m18_var = tk.BooleanVar(value=OUT_ZIGZAG)
    m19_var = tk.BooleanVar(value=OUT_SCORE)

    vars_map = {}          # 板块名 -> BooleanVar（跨类别常驻，切换类别不丢勾选）

    # —— 顶部：类别单选 + 搜索 ——
    top = ttk.Frame(root, padding=(12, 10, 12, 4))
    top.pack(fill='x')
    ttk.Label(top, text="板块类别：").pack(side='left')
    for c in ('概念', '行业', '全部'):
        ttk.Radiobutton(top, text=c, value=c, variable=cat_var,
                        command=lambda: switch_category()).pack(side='left', padx=6)
    ttk.Label(top, text="   搜索：").pack(side='left')
    ent = ttk.Entry(top, textvariable=kw_var, width=18)
    ent.pack(side='left')
    kw_var.trace_add('write', lambda *a: refresh_list())

    # —— 中部：板块复选框（随窗口宽度自适应列数）——
    box = ttk.Frame(root, padding=(12, 4))
    box.pack(fill='both', expand=True)
    canvas = tk.Canvas(box, highlightthickness=0)
    bar = ttk.Scrollbar(box, orient='vertical', command=canvas.yview)
    inner = ttk.Frame(canvas)
    inner.bind('<Configure>', lambda e: canvas.configure(scrollregion=canvas.bbox('all')))
    win_id = canvas.create_window((0, 0), window=inner, anchor='nw')
    canvas.configure(yscrollcommand=bar.set)
    canvas.pack(side='left', fill='both', expand=True)
    bar.pack(side='right', fill='y')
    canvas.bind_all('<MouseWheel>', lambda e: canvas.yview_scroll(int(-e.delta / 120), 'units'))

    COL_W = 210
    layout = {'cols': 0}

    def apply_grid(width):
        cols = layout['cols'] or 4
        cw = max(140, (width - 16) // cols)
        for c in range(12):
            inner.grid_columnconfigure(c, weight=0, minsize=cw if c < cols else 0)

    def on_canvas_resize(e):
        canvas.itemconfigure(win_id, width=e.width)
        cols = max(1, min(10, e.width // COL_W))
        if cols != layout['cols']:
            layout['cols'] = cols
            refresh_list()
        else:
            apply_grid(e.width)
    canvas.bind('<Configure>', on_canvas_resize)

    # —— 底部：统计文字 + 三行控件 ——
    bottom = ttk.Frame(root, padding=(12, 6, 12, 10))
    bottom.pack(fill='x')
    tip = ttk.Label(bottom, text="", anchor='w', justify='left')
    tip.pack(fill='x', pady=(0, 6))
    opts = ttk.Frame(bottom);  opts.pack(fill='x', pady=(0, 6))
    outs = ttk.Frame(bottom);  outs.pack(fill='x', pady=(0, 6))
    acts = ttk.Frame(bottom);  acts.pack(fill='x')

    def on_root_resize(e=None):
        w = root.winfo_width()
        if w > 100 and abs(w - layout.get('w', 0)) > 20:
            layout['w'] = w
            tip.config(wraplength=w - 40)
    root.bind('<Configure>', on_root_resize)

    def current_items():
        if cat_var.get() == '全部':
            items = []
            for c in ('概念', '行业'):
                items += [(n, k, c) for n, k in catalog.get(c, [])]
            for c in catalog:
                if c not in ('概念', '行业'):
                    items += [(n, k, c) for n, k in catalog[c]]
        else:
            items = [(n, k, cat_var.get()) for n, k in catalog.get(cat_var.get(), [])]
        kw = kw_var.get().strip().lower()
        if kw:
            items = [i for i in items if kw in i[0].lower()]
        return items

    def selected():
        return [n for n, v in vars_map.items() if v.get()]

    def update_tip(*a):
        try:
            n = int(topn_var.get() or 0)
        except ValueError:
            n = 0
        sel = selected()
        visible = {nm for nm, _, _ in current_items()}
        other = [x for x in sel if x not in visible]
        extra = f"｜其中 {len(other)} 个不在当前视图：{'、'.join(other[:6])}" if other else ""
        tip.config(text=f"已选 {len(sel)} 个板块 ≈ {P.preview_count(pool, sel, n)} 只股票（已按代码去重）{extra}")

    def refresh_list():
        for w in inner.winfo_children():
            w.destroy()
        items = current_items()
        cols = layout['cols'] or 4
        for i, (nm, cnt, cat) in enumerate(items):
            if nm not in vars_map:
                vars_map[nm] = tk.BooleanVar(value=False)
            txt = f"{nm}（{cnt}）" if cat_var.get() != '全部' else f"{nm}·{cat}（{cnt}）"
            cb = ttk.Checkbutton(inner, text=txt, variable=vars_map[nm], command=update_tip)
            cb.grid(row=i // cols, column=i % cols, sticky='w', padx=6, pady=3)
        apply_grid(canvas.winfo_width() or 900)
        update_tip()

    def switch_category():
        refresh_list()

    def check_all(flag):
        for nm, _, _ in current_items():
            vars_map.setdefault(nm, tk.BooleanVar(value=False)).set(flag)
        update_tip()

    def restore_default():
        for v in vars_map.values():
            v.set(False)
        for nm in P._defaults_for(catalog, '全部', defaults=_default_boards()):
            vars_map.setdefault(nm, tk.BooleanVar(value=False)).set(True)
        topn_var.set(str(DEFAULT_TOP_N_PER_BOARD))
        update_tip()

    def set_topn(n):
        topn_var.set(str(n))
        update_tip()

    def confirm():
        sel = selected()
        if not sel and not broad_var.get() and not ovs_var.get():
            messagebox.showwarning("提示", "请至少勾选一个板块（或勾上宽基指数 / 国外标的）")
            return
        try:
            n = max(0, int(topn_var.get() or 0))
        except ValueError:
            n = 0
        if not any([charts_var.get(), excel_var.get(), m17_var.get(), m18_var.get(), m19_var.get()]):
            messagebox.showwarning("提示", "请至少勾选一项输出内容")
            return
        result.update(boards=sel, top_n=n, category=cat_var.get(),
                      overseas=ovs_var.get(), legacy=legacy_var.get(), broad=broad_var.get(),
                      charts=charts_var.get(), excel=excel_var.get(),
                      m17=m17_var.get(), m18=m18_var.get(), m19=m19_var.get())
        root.destroy()

    # —— 第一行：选择辅助 + 取数范围 + Top N 快捷按钮 ——
    ttk.Button(opts, text="恢复默认", command=restore_default).pack(side='left', padx=(0, 4))
    ttk.Button(opts, text="全选", command=lambda: check_all(True)).pack(side='left', padx=4)
    ttk.Button(opts, text="清空", command=lambda: check_all(False)).pack(side='left', padx=4)
    ttk.Label(opts, text="   每板块取前 N 只：").pack(side='left', padx=(12, 4))
    ttk.Entry(opts, textvariable=topn_var, width=5).pack(side='left')
    for n in TOP_N_PRESETS:
        label = "全部" if n == 0 else f"前{n}"
        ttk.Button(opts, text=label, width=5,
                   command=lambda v=n: set_topn(v)).pack(side='left', padx=2)
    topn_var.trace_add('write', update_tip)

    # —— 第二行：并入范围 + 输出模块 ——
    ttk.Label(outs, text="输出：").pack(side='left')
    ttk.Checkbutton(outs, text="① K线形态图", variable=charts_var).pack(side='left', padx=4)
    ttk.Checkbutton(outs, text="① 板块统计Excel", variable=excel_var).pack(side='left', padx=4)
    ttk.Checkbutton(outs, text="② 八大指标矩阵", variable=m17_var).pack(side='left', padx=4)
    ttk.Checkbutton(outs, text="③ 波段信号看板", variable=m18_var).pack(side='left', padx=4)
    ttk.Checkbutton(outs, text="④ 评分矩阵+T+1", variable=m19_var).pack(side='left', padx=4)
    ttk.Checkbutton(outs, text="并入 target_stocks 国内", variable=legacy_var).pack(side='right', padx=(8, 0))
    ttk.Checkbutton(outs, text="国外标的", variable=ovs_var).pack(side='right', padx=6)
    ttk.Checkbutton(outs, text="宽基指数", variable=broad_var).pack(side='right', padx=6)

    # —— 第三行：说明 + 主操作按钮 ——
    ttk.Label(acts, text="N 按「基金持有比例(%)」从高到低排序，0 = 全部",
              foreground="#666").pack(side='left')
    ttk.Button(acts, text="取消", command=root.destroy).pack(side='right', padx=(4, 0))
    ttk.Button(acts, text="开始分析", command=confirm).pack(side='right', padx=4)

    restore_default()
    refresh_list()
    root.mainloop()
    return result or None


def select_options_console(pool, catalog):
    """无图形界面时的命令行选择：直接回车 = 用默认"""
    P = pm()
    choice = _base_choice()

    print("\n请选择板块类别： 1) 概念(默认)  2) 行业  3) 全部")
    ans = input("输入序号后回车（直接回车用默认）: ").strip()
    category = {'2': '行业', '3': '全部'}.get(ans, DEFAULT_BOARD_CATEGORY)

    items = []
    if category == '全部':
        for c in catalog:
            items += [(n, k) for n, k in catalog[c]]
    else:
        items = catalog.get(category, [])

    default_sel = P._defaults_for(catalog, '全部', defaults=_default_boards())
    print(f"\n【{category}】共 {len(items)} 个板块：")
    for i, (n, k) in enumerate(items, 1):
        mark = '*' if n in default_sel else ' '
        print(f"  {mark}{i:>3}. {n}({k})", end='\n' if i % 4 == 0 else '')
    print(f"\n（带 * 为默认勾选：{'、'.join(default_sel)}）")

    raw = input("输入板块序号或名称，逗号分隔（直接回车用默认）: ").strip()
    if not raw:
        boards = default_sel
    else:
        boards, names = [], [n for n, _ in items]
        for p in raw.replace('，', ',').split(','):
            p = p.strip()
            if not p:
                continue
            if p.isdigit() and 1 <= int(p) <= len(names):
                boards.append(names[int(p) - 1])
            elif p in names:
                boards.append(p)
        boards = list(dict.fromkeys(boards)) or default_sel

    raw_n = input(f"每板块取前 N 只(按基金持有比例, 0=全部, 默认 {DEFAULT_TOP_N_PER_BOARD}): ").strip()
    try:
        top_n = max(0, int(raw_n)) if raw_n else DEFAULT_TOP_N_PER_BOARD
    except ValueError:
        top_n = DEFAULT_TOP_N_PER_BOARD

    raw_m = input("输出模块（1形态图 2Excel 3指标矩阵 4波段看板 5评分矩阵，逗号分隔，回车=全部）: ").strip()
    if raw_m:
        picked = {p.strip() for p in raw_m.replace('，', ',').split(',')}
        choice.update(charts='1' in picked, excel='2' in picked,
                      m17='3' in picked, m18='4' in picked, m19='5' in picked)

    choice.update(boards=boards, top_n=top_n, category=category)
    return choice


def choose_options(pool):
    """统一入口：优先图形勾选，失败退回命令行，再失败用配置默认值"""
    if BROAD_INDEX_ONLY:
        # 开源版：跳过板块勾选，直接用配置默认值（boards=[]、broad=True），只出宽基指数。
        return _base_choice()

    P = pm()
    catalog = P.board_catalog(pool)
    choice = _base_choice()
    if not catalog:
        return choice

    if USE_SELECTION_UI:
        try:
            res = select_options_ui(pool, catalog)
        except Exception as e:
            print(f"⚠️ 勾选窗口异常({e})，改用命令行选择")
            res = None
        if res:
            choice.update(res)
            return choice
        print("（未检测到图形界面或窗口被关闭，改用命令行选择）")
        try:
            return select_options_console(pool, catalog)
        except (EOFError, KeyboardInterrupt):
            pass

    choice['boards'] = P._defaults_for(catalog, '全部', defaults=_default_boards())
    return choice


# ==========================================================================
# 4. 标的组装（一只票横跨多个板块时按代码去重：只抓一次数、只画一张图）
# ==========================================================================
def build_tasks(pool, choice):
    P = pm()
    boards, top_n = choice['boards'], choice['top_n']
    name2cat = dict(zip(pool['板块'], pool['类别'])) if not pool.empty else {}

    dom = P.build_domestic_targets(pool, boards, top_n) if boards else {}
    legacy = P.build_legacy_targets(LEGACY_CSV, name2cat,
                                    include_overseas=choice['overseas'],
                                    include_domestic=choice['legacy'])
    targets = P.merge_targets(dom, legacy)

    if choice['broad']:
        for code, info in P.BROAD_INDICES.items():
            targets[code] = dict(code=code, name=info[0], region=info[3], kind='宽基',
                                 boards=[{'板块': info[4], '类别': '宽基'}])

    # 「基金持有比例」：一只票在多个板块里取最大值，用于 ZigZag 看板的截断排序
    ratio_map = {}
    if boards:
        sub = pool[pool['板块'].isin(boards)]
        if top_n and top_n > 0:
            sub = (sub.sort_values('ratio', ascending=False, na_position='last')
                      .groupby('板块').head(top_n))
        for c, g in sub.groupby('code'):
            v = pd.to_numeric(g['ratio'], errors='coerce')
            ratio_map[c] = float(v.max()) if v.notna().any() else np.nan

    tasks = []
    for t in targets.values():
        bs = t['boards'] or [{'板块': '', '类别': '未分类'}]
        raw_code = t['code']
        code = P.normalize_code(raw_code)
        cats = list(dict.fromkeys([b.get('类别') or '未分类' for b in bs]))
        tasks.append(dict(
            code=code,
            raw_code=raw_code,
            name=t['name'],
            disp=f"{t['name']}({short_code(code)})",
            key=safe_key(code),
            sector=bs[0]['板块'],                                            # 主板块：文件名前缀
            sector_display=' / '.join(b['板块'] for b in bs if b['板块']),   # 标题展示全部板块
            boards=bs,
            board_names=[b['板块'] for b in bs if b['板块']],
            cats=cats,
            region=t.get('region', '国内'),
            kind=t.get('kind', '板块'),
            ratio=ratio_map.get(raw_code, np.nan),
            zigzag=False,
        ))

    # ZigZag 看板会把每只票的两张图内嵌进同一页，标的多了页面会打不开 —— 这里按持仓比例截断。
    # 宽基指数数量少且参考价值高，排序时优先保留。
    if choice['m18']:
        # 上限只约束个股：宽基指数总共十来个、参考价值高，不占个股名额（
        # ZIGZAG_INDEX_ALWAYS 置 False 则一并参与排序竞争）。
        idx_tasks = [t for t in tasks if t['kind'] == '宽基']
        stk_tasks = [t for t in tasks if t['kind'] != '宽基']
        if not ZIGZAG_INDEX_ALWAYS:
            stk_tasks = stk_tasks + idx_tasks
            idx_tasks = []
        # 个股按「基金持有比例」从高到低，缺值的排最后
        stk_tasks.sort(key=lambda t: -(t['ratio'] if pd.notna(t['ratio']) else -1))
        keep = stk_tasks if not ZIGZAG_MAX_TARGETS else stk_tasks[:ZIGZAG_MAX_TARGETS]
        for t in keep + idx_tasks:
            t['zigzag'] = True
        if ZIGZAG_MAX_TARGETS and len(stk_tasks) > ZIGZAG_MAX_TARGETS:
            print(f"ℹ️ 波段看板只收录基金持仓比例最高的 {ZIGZAG_MAX_TARGETS} 只个股"
                  f"（共 {len(stk_tasks)} 只，其余仍会进入指标矩阵与 Excel）"
                  + ("，宽基指数不占名额" if idx_tasks else ""))
    return tasks


# ==========================================================================
# 5. 单只标的的完整分析（放在进程池里跑）
# ==========================================================================
def _relabel_nav_fig(fig):
    """18 号内核的图是给基金画的，把「净值」字样换成个股口径的「价格」。"""
    for tr in fig.data:
        if getattr(tr, 'name', None):
            tr.name = str(tr.name).replace('单位净值', '收盘价').replace('净值', '价格')
        ht = getattr(tr, 'hovertemplate', None)
        if ht:
            tr.hovertemplate = str(ht).replace('单位净值', '收盘价').replace('净值', '价格')
    try:
        title = fig.layout.title.text or ''
        fig.update_layout(title_text=title.replace('历史净值已复权', '前复权'),
                          yaxis_title='收盘价(前复权)')
    except Exception:
        fig.update_layout(yaxis_title='收盘价(前复权)')
    return fig


def analyze_one(args):
    """【进程池】形态识别 + 绘图 + 八大指标 + 过热评分 + T+1 模拟 + ZigZag 图表。
    只把体量很小的结果传回主进程。"""
    task, df, cfg = args
    res = dict(task=task, err=None, signal_df=None, ind_tail=None,
               heat=None, sims=None, zz=None)
    try:
        P, C = pm(), core()
        P.SIGNAL_RECENT_DAYS = cfg['signal_days']
        P.PLOT_RECENT_BARS = cfg['plot_bars']
        P.PLOTLY_JS_MODE = cfg['plotly_js']

        # ---------- ① 形态 / 缺口 / 量价 / 涨跌停 ----------
        if cfg['charts'] or cfg['excel']:
            dfp = P.extract_pattern_names(df)
            if cfg['excel']:
                res['signal_df'] = P.build_signal_frame(
                    dfp, task['kind'], task['region'], task['boards'], task['code'])
            if cfg['charts']:
                P.plot_interactive(dfp, task['name'], task['code'], task['sector'],
                                   cfg['chart_dir'], sector_display=task['sector_display'])

        if not (cfg['m17'] or cfg['m18'] or cfg['m19']):
            return res

        # ---------- ② 八大指标 + 过热评分（口径 = 18 号内核）----------
        # 个股取的是前复权价，本身连续可比，故不走基金那套 build_adjusted_nav 复权
        close = pd.to_numeric(df['Close'], errors='coerce').dropna()
        close = close[~close.index.duplicated(keep='last')].sort_index()
        if cfg['max_bars'] and len(close) > cfg['max_bars']:
            close = close.tail(cfg['max_bars'])
        if len(close) < C.MIN_ROWS:
            res['err'] = f"⚠️ {task['name']} ({task['code']}) 有效数据不足 {C.MIN_ROWS} 条，指标部分跳过"
            return res

        ind = C.compute_eight_indicators(close)
        net_s, int_s, all_red_s, all_green_s = C.compute_heat_score_series(ind)
        score_s = C.calc_score_series(net_s, int_s)

        df_ind = pd.DataFrame({
            'macd_hist': ind['macd'],
            'rsi': ind['rsi'],
            'bias': ind['bias'],
            'kdj': ind['kdj'],
            'cci': ind['cci'],
            'sar_stat': np.where(ind['sar_bullish'].fillna(False).astype(bool), '看涨', '看跌'),
            'boll_pctb': ind['boll_pctb'],
            'boll_bw_pct': ind['bw_pct'],
            'score': score_s.reindex(ind.index),
        }, index=ind.index)
        res['ind_tail'] = df_ind.tail(cfg['hist_days']).copy()

        latest_net, latest_int = net_s.iloc[-1], int_s.iloc[-1]
        latest_score = score_s.iloc[-1]
        score_val = None if pd.isna(latest_score) else float(latest_score)
        status, meaning, action, color = C.get_status_action(score_val)
        res['heat'] = {
            'net_count': -99 if pd.isna(latest_net) else float(latest_net),
            'intensity': -99 if pd.isna(latest_int) else float(latest_int),
            'score': 0.0 if score_val is None else score_val,
            'status': status, 'meaning': meaning, 'action': action, 'color': color,
            'last_close': float(close.iloc[-1]),
            'last_date': close.index[-1].strftime('%Y-%m-%d'),
        }

        # ---------- ③ T+1 涨跌情景模拟 ----------
        if cfg['m19']:
            sims = {}
            for pct in cfg['sim_returns']:
                ind_row, sim_score = simulate_next_day(close, pct)
                st, mn, ac, cl = C.get_status_action(sim_score)
                delta = None
                if sim_score is not None and score_val is not None:
                    delta = sim_score - score_val
                sims[pct] = {'score': sim_score, 'status': st, 'meaning': mn, 'action': ac,
                             'color': cl, 'delta': delta,
                             'tip': _indicator_tip(ind_row) if ind_row else ''}
            res['sims'] = sims

        # ---------- ④ ZigZag 波浪 + 波段信号图表 ----------
        if cfg['m18'] and task['zigzag']:
            df_plot = close.to_frame('close')
            df_plot.index.name = '日期'
            df_zig = C.calculate_zigzag(df_plot, price_col='close', change_pct=C.ZIGZAG_THRESHOLD)
            pivots_df = df_zig[df_zig['pivot'] != 0].copy()
            peaks = pivots_df[pivots_df['pivot'] == 1]
            troughs = pivots_df[pivots_df['pivot'] == -1]

            last_peak_val = float(peaks['close'].iloc[-1]) if not peaks.empty else 'null'
            last_trough_val = float(troughs['close'].iloc[-1]) if not troughs.empty else 'null'
            last_date = df_plot.index[-1]
            first_date = df_plot.index[0]
            last_close = float(df_plot['close'].iloc[-1])
            next_date = last_date + pd.tseries.offsets.BDay(1)

            fixed_indicators, pct_snapshot, signal, ind_history = C.compute_fund_signal(df_plot['close'])

            key = task['key']
            nav_fig = _relabel_nav_fig(C.build_nav_figure(
                df_plot, pivots_df, peaks, troughs, short_code(task['code']), task['name'],
                last_date, last_close, next_date, last_peak_val, last_trough_val,
                split_events=None,
            ))
            nav_html = nav_fig.to_html(full_html=False, include_plotlyjs=False,
                                       div_id=f"nav-chart-{key}", config=C.PLOTLY_CONFIG)
            score_fig = C.build_score_figure(task['name'], short_code(task['code']),
                                             signal, ind_history)
            score_html = score_fig.to_html(full_html=False, include_plotlyjs=False,
                                           div_id=f"score-chart-{key}", config=C.PLOTLY_CONFIG)

            res['zz'] = {
                'key': key,
                'nav_html': f'<div class="nav-chart-wrap" id="nav-wrap-{key}">{nav_html}</div>',
                'score_html': f'<div class="score-chart-wrap" id="score-wrap-{key}">{score_html}</div>',
                'meta': {
                    # 页面 JS 会自己拼「名称（代码）」，这里只放名称，避免代码出现两遍
                    'name': task['name'],
                    'last_nav': round(last_close, 4),
                    'peak': None if last_peak_val == 'null' else round(float(last_peak_val), 4),
                    'trough': None if last_trough_val == 'null' else round(float(last_trough_val), 4),
                    'last_date': last_date.strftime('%Y-%m-%d'),
                    'first_date': first_date.strftime('%Y-%m-%d'),
                },
                'fixed': C.style_fixed_indicators(fixed_indicators),
                'pct': {w: C.style_percentile(pct_snapshot[w]) for w in C.WINDOW_ORDER},
                'signal_latest': {'latest': signal['latest']},
                'score_history': signal['score_history'],
            }
        return res
    except Exception as e:
        res['err'] = f"❌ {task.get('name')} ({task.get('code')}) 分析失败: {e}"
        return res


# ==========================================================================
# 6. 板块形态统计汇总 Excel（口径与原 Price_movement_patterns.py 完全一致）
# ==========================================================================
def write_pattern_excel(all_signals_list, output_directory):
    import openpyxl
    from openpyxl.styles import PatternFill

    P = pm()
    if not all_signals_list:
        print("\n❌ 没有数据用于生成汇总报表。")
        return

    print("\n>>> 正在聚合全市场信号，生成 Excel 汇总表...")
    combined_df = pd.concat(all_signals_list)
    combined_df.reset_index(inplace=True)

    if 'Date' in combined_df.columns:
        combined_df.rename(columns={'Date': '日期'}, inplace=True)
    elif 'index' in combined_df.columns:
        combined_df.rename(columns={'index': '日期'}, inplace=True)

    combined_df['日期'] = pd.to_datetime(combined_df['日期']).dt.strftime('%Y-%m-%d')

    METRICS = ['看涨', '看跌', '向上缺口', '向下缺口', '量价看多', '量价看空', '涨停', '跌停']

    def build_agg(df_src, keys):
        """按给定维度聚合：数量去重计数 + 各项信号求和 + 重算利好/利空/多空比"""
        cnt = df_src.groupby(keys)['股票代码'].nunique().reset_index(name='板块数量')
        summ = df_src.groupby(keys)[METRICS].sum().reset_index()
        out = pd.merge(summ, cnt, on=keys)

        if '收益率' in df_src.columns:
            ret = df_src.groupby(keys)['收益率'].mean().reset_index()
            ret.rename(columns={'收益率': '板块收益'}, inplace=True)
            out = pd.merge(out, ret, on=keys, how='left')
        else:
            out['板块收益'] = np.nan

        if P.SIGNAL_SCHEME == 'v1':
            out['利好'] = out['看涨'] + out['向上缺口'] + out['量价看多']
            out['利空'] = out['看跌'] + out['向下缺口'] + out['量价看空']
        else:
            out['利好'] = out['向上缺口'] + out['量价看多']
            out['利空'] = out['向下缺口'] + out['量价看空']
        out['多空指标比'] = ((out['利好'] - out['利空'])
                             / out['板块数量'].replace(0, np.nan)).round(4)
        out['涨跌停净'] = out['涨停'] - out['跌停']
        out['涨停率'] = (out['涨停'] / out['板块数量'].replace(0, np.nan)).round(4)

        cols = (keys + ['板块数量'] + METRICS
                + ['利好', '利空', '多空指标比', '涨跌停净', '涨停率', '板块收益'])
        out = out[cols]
        out.sort_values(by=['日期'] + keys[1:],
                        ascending=[False] + [True] * (len(keys) - 1), inplace=True)
        return out

    if '板块类别' not in combined_df.columns:
        combined_df['板块类别'] = '未分类'

    detail_df = build_agg(combined_df, ['日期', '类型', '归属', '板块类别', '板块'])
    global_df = build_agg(combined_df, ['日期', '类型', '板块类别', '板块'])
    global_df.insert(2, '归属', '全球合并')
    dom_df = build_agg(combined_df[combined_df['归属'] == '国内'],
                       ['日期', '类型', '板块类别', '板块'])
    ovs_src = combined_df[combined_df['归属'] == '国外']
    ovs_df = build_agg(ovs_src, ['日期', '类型', '板块类别', '板块']) if not ovs_src.empty else pd.DataFrame()

    # ---- 市场基准：按去重个股统计全市场等权收益与净广度，供 7.py 剔 beta 用 ----
    uni = (combined_df[combined_df['类型'] == '板块']
           .drop_duplicates(['日期', '归属', '股票代码']))
    mkt = uni.groupby(['日期', '归属']).agg(
        全市场样本数=('股票代码', 'nunique'),
        全市场收益=('收益率', 'mean'),
        全市场利好=('量价看多', 'sum'),
        全市场利空=('量价看空', 'sum'),
        全市场涨停=('涨停', 'sum'),
        全市场跌停=('跌停', 'sum'),
    ).reset_index()
    mkt['全市场净广度'] = ((mkt['全市场利好'] - mkt['全市场利空'])
                          / mkt['全市场样本数'].replace(0, np.nan)).round(4)
    mkt.sort_values(['日期', '归属'], ascending=[False, True], inplace=True)

    meta = pd.DataFrame([
        {"键": "SIGNAL_SCHEME", "值": P.SIGNAL_SCHEME},
        {"键": "利好口径", "值": ("看涨+向上缺口+量价看多" if P.SIGNAL_SCHEME == 'v1'
                                  else "向上缺口+量价看多")},
        {"键": "利空口径", "值": ("看跌+向下缺口+量价看空" if P.SIGNAL_SCHEME == 'v1'
                                  else "向下缺口+量价看空")},
        {"键": "涨跌停口径", "值": ("收盘封板" if P.LIMIT_SEAL_ONLY else "含炸板")
                                   + "；主板10%/创业板科创板20%/北交所30%/主板ST5%；境外标的不统计"},
        {"键": "SIGNAL_RECENT_DAYS", "值": P.SIGNAL_RECENT_DAYS},
        {"键": "生成时间", "值": time.strftime("%Y-%m-%d %H:%M:%S")},
    ])

    excel_path = os.path.join(output_directory, "2.板块形态统计汇总.xlsx")
    sheets = {
        "元信息": meta,
        "明细": detail_df,          # 供 Sector_heatmap.py 读取
        "全球合并": global_df,
        "国内": dom_df,
        "国外": ovs_df,
        "行业": global_df[global_df['板块类别'] == '行业'],
        "概念": global_df[global_df['板块类别'] == '概念'],
        "宽基": detail_df[detail_df['类型'] == '宽基'],
        "市场基准": mkt,
    }
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        for sname, sdf in sheets.items():
            if sdf is not None and not sdf.empty:
                sdf.to_excel(writer, sheet_name=sname, index=False)

    try:
        wb = openpyxl.load_workbook(excel_path)
        red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
        green_fill = PatternFill(start_color="99FF99", end_color="99FF99", fill_type="solid")

        for ws in wb.worksheets:
            headers = [cell.value for cell in ws[1]]

            if '利好' in headers and '利空' in headers:
                lihao_idx = headers.index('利好') + 1
                likong_idx = headers.index('利空') + 1
                for row in range(2, ws.max_row + 1):
                    lihao_val = ws.cell(row=row, column=lihao_idx).value
                    if isinstance(lihao_val, (int, float)) and lihao_val > 0:
                        ws.cell(row=row, column=lihao_idx).fill = red_fill
                    likong_val = ws.cell(row=row, column=likong_idx).value
                    if isinstance(likong_val, (int, float)) and likong_val > 0:
                        ws.cell(row=row, column=likong_idx).fill = green_fill

            if '多空指标比' in headers:
                ratio_col_idx = headers.index('多空指标比') + 1
                for row in range(2, ws.max_row + 1):
                    val = ws.cell(row=row, column=ratio_col_idx).value
                    if isinstance(val, (int, float)):
                        if val > 0.2:
                            ws.cell(row=row, column=ratio_col_idx).fill = red_fill
                        elif val < -0.2:
                            ws.cell(row=row, column=ratio_col_idx).fill = green_fill

            ws.freeze_panes = "A2"
            for col_i, h in enumerate(headers, start=1):
                width = max(10, min(16, len(str(h)) * 2 + 4))
                ws.column_dimensions[ws.cell(row=1, column=col_i).column_letter].width = width

        wb.save(excel_path)
        print(f"✅ 成功生成带条件格式的汇总报表: {excel_path}")
        print(f"   工作表: {' / '.join(k for k, v in sheets.items() if v is not None and not v.empty)}")
    except Exception as e:
        print(f"✅ 成功生成报表，但单元格染色应用失败: {e}")


# ==========================================================================
# 7. 页面公共零件（汇总条 / 分类按钮 / 板块下拉）
# ==========================================================================
def collect_rows(results):
    """把并发返回的结果整理成矩阵页需要的行，排序口径与基金版一致：
    先按 net_count 降序（全红在前、全绿在后），同 net_count 再按 intensity 降序。"""
    rows = []
    for r in results:
        if r.get('ind_tail') is None or r.get('heat') is None:
            continue
        t = r['task']
        rows.append(dict(
            disp=t['disp'], name=t['name'], code=t['code'], key=t['key'],
            cats=t['cats'], boards=t['board_names'],
            bpairs=[(b.get('板块') or '', b.get('类别') or '未分类') for b in t['boards']],
            kind=t['kind'], region=t['region'], ratio=t['ratio'],
            ind=r['ind_tail'], heat=r['heat'], sims=r.get('sims'), zz=r.get('zz'),
        ))
    rows.sort(key=lambda x: (x['heat']['net_count'], x['heat']['intensity']), reverse=True)
    return rows


def _summary_bar_html(rows):
    counts = {s: 0 for s in STATUS_ORDER}
    for r in rows:
        st = r['heat'].get('status', '⚪ 无数据')
        counts[st if st in counts else '⚪ 无数据'] += 1

    html = '<div class="summary-bar">'
    for st in STATUS_ORDER:
        color = STATUS_COLOR_MAP[st]
        slug = STATUS_SLUG_MAP[st]
        cnt = counts[st]
        style = 'display:none;' if cnt == 0 else ''
        html += (f'<div class="summary-item" data-status="{slug}" style="border-color:{color};{style}">'
                 f'<span class="summary-label" style="color:{color};">{st}</span>'
                 f'<span class="summary-count">{cnt}</span></div>')
    html += (f'<div class="summary-total">共 <span id="summary-total-num">{len(rows)}</span> 只标的</div>'
             '</div>')
    return html


def _filter_bar_html(rows):
    """具体板块下拉筛选（比如 A股宽基 / 港股宽基 / 美股宽基 / 外盘宽基）。
    BROAD_INDEX_ONLY 模式下所有标的都是「宽基」这一个类别，类别切换按钮没有意义，
    只保留板块下拉。"""
    cats, cat2boards = [], {}
    for r in rows:
        for b, c in r['bpairs']:
            if c not in cats:
                cats.append(c)
            if b:
                cat2boards.setdefault(c, set()).add(b)

    opts = '<option value="">全部板块</option>'
    for c in cats:
        bs = sorted(cat2boards.get(c, []))
        if not bs:
            continue
        opts += f'<optgroup label="{_esc(c)}">'
        for b in bs:
            n = sum(1 for r in rows if b in r['boards'])
            opts += f'<option value="{_esc(b)}">{_esc(b)}（{n}）</option>'
        opts += '</optgroup>'

    return (
        '<div class="btn-group filter-row">'
        '<span style="margin-right:10px; color:#aaa;">板块：</span>'
        f'<select id="board-select" class="board-select">{opts}</select>'
        '</div>'
    )


def _row_head_cell(r):
    """矩阵页最左侧的标的单元格：名称 + 所属板块 + 风险徽章"""
    hs = r['heat']
    status = hs.get('status', '⚪ 无数据')
    color = hs.get('color', '#888888')
    score = hs.get('score', 0)
    action = _esc(hs.get('action', ''))
    boards = r['boards']
    btxt = ' / '.join(boards[:3]) + ('…' if len(boards) > 3 else '')
    ratio = r.get('ratio')
    rtxt = '' if ratio is None or pd.isna(ratio) else f'　基金持股 {ratio:.2f}%'
    return (
        f'<td class="index-name-td">'
        f'<div class="fund-name">{_esc(r["disp"])}</div>'
        f'<div class="stk-boards" title="{_esc(" / ".join(boards))}{_esc(rtxt)}">{_esc(btxt)}</div>'
        f'<div class="risk-badge" style="color:{color};border-color:{color};" title="{action}">'
        f'{status} · {score:.0f}分</div>'
        f'</td>'
    )


# ==========================================================================
# 8. 矩阵页公共 CSS / JS（17 与 19 共用一套，改样式只需改一处）
# ==========================================================================
MATRIX_CSS = r"""
        body {
            background-color: #161616;
            color: #ffffff;
            font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, "Helvetica Neue", Arial, sans-serif;
            margin: 0;
            padding: 20px;
        }
        h2 { text-align: center; color: #E5C07B; margin-bottom: 6px; }
        .subtitle { text-align: center; color: #777; font-size: 12px; margin-bottom: 12px; }
        .btn-group { margin: 10px 0; text-align: center; }
        .btn-group.indicator-row, .btn-group.tool-row {
            display: flex; align-items: center; justify-content: space-between;
            flex-wrap: wrap; gap: 10px; text-align: left;
        }
        .indicator-row-left, .tool-row-left {
            display: flex; align-items: center; flex-wrap: wrap; text-align: left;
        }
        .btn-group.filter-row {
            display: flex; align-items: center; justify-content: center; flex-wrap: wrap;
        }
        .summary-bar {
            display: flex; flex-wrap: wrap; justify-content: center; align-items: center;
            gap: 10px; margin: 10px 0 15px 0;
        }
        .summary-item {
            display: flex; align-items: center; gap: 6px; background-color: #1e1e1e;
            border: 1px solid; border-radius: 20px; padding: 5px 14px; font-size: 12px;
        }
        .summary-label { font-weight: bold; }
        .summary-count {
            color: #fff; font-weight: bold; background-color: rgba(255,255,255,0.1);
            border-radius: 10px; padding: 1px 8px;
        }
        .summary-total { color: #888; font-size: 12px; margin-left: 8px; }
        .btn-group button {
            background-color: #333; color: #ccc; border: 1px solid #555;
            padding: 6px 14px; margin: 0 4px 8px 4px; border-radius: 4px;
            cursor: pointer; font-size: 12px; transition: background-color 0.2s, color 0.2s;
        }
        .btn-group button.active { background-color: #61AFEF; color: #fff; border-color: #61AFEF; }
        .btn-group button:hover { background-color: #444; }
        .btn-group button.active:hover { background-color: #528BC6; }
        .board-select {
            background-color: #333; color: #ccc; border: 1px solid #555;
            border-radius: 4px; padding: 6px 10px; font-size: 12px; max-width: 260px;
        }
        .table-container {
            overflow-x: auto; max-width: 100%; border: 1px solid #333; border-radius: 8px;
            background-color: #1e1e1e; box-shadow: 0 4px 12px rgba(0,0,0,0.5);
        }
        table { border-collapse: collapse; width: 100%; font-size: 11px; white-space: nowrap; }
        th, td { border: 1px solid #2a2a2a; padding: 4px 6px; text-align: center; vertical-align: middle; }
        th {
            background-color: #252525; color: #61AFEF; position: sticky; top: 0;
            z-index: 10; font-size: 12px; padding: 8px 10px;
        }
        th.sim-col { background-color: #1d2735; color: #7FB3FF; }
        td.sim-col { background-color: rgba(97,175,239,0.05); }
        th.sim-first, td.sim-first { border-left: 2px solid #3a5a80; }
        th.sim-last,  td.sim-last  { border-right: 2px solid #3a5a80; }
        .sim-head-main { display:block; font-weight:bold; }
        .sim-head-sub  { display:block; font-size:10px; color:#6b7f99; font-weight:normal; }
        .index-name-th {
            position: sticky; left: 0; background-color: #252525; z-index: 20;
            font-weight: bold; color: #C678DD; text-align: left; padding-left: 12px;
        }
        .index-name-td {
            position: sticky; left: 0; background-color: #1b1b1b; z-index: 5;
            font-weight: bold; color: #C678DD; text-align: left; padding-left: 12px;
            vertical-align: top; padding-top: 10px;
        }
        .indicators-grid {
            display: grid; grid-template-columns: repeat(2, 1fr); gap: 2px;
            font-size: 10px; text-align: left; min-width: 130px;
        }
        .indicator-item {
            background: rgba(255, 255, 255, 0.02); padding: 2px 4px; border-radius: 2px;
            display: flex; justify-content: space-between;
        }
        .ind-label { color: #888; margin-right: 3px; }
        .ind-val { font-weight: bold; }
        .score-badge {
            display: inline-flex; align-items: center; gap: 5px; border: 1px solid;
            border-radius: 12px; padding: 3px 9px; font-size: 11px; line-height: 1.2;
            white-space: nowrap; cursor: default;
        }
        .score-status { font-weight: normal; }
        .score-num { font-weight: bold; }
        .score-delta { font-size: 10px; font-weight: bold; }
        .score-badge.sim-badge { box-shadow: 0 0 0 1px rgba(255,255,255,0.04) inset; }
        .hidden { display: none; }
        .col-hidden { display: none; }
        .fund-name { margin-bottom: 3px; }
        .stk-boards { font-size: 10px; color: #7a7a7a; font-weight: normal; margin-bottom: 4px; }
        .risk-badge {
            display: inline-block; font-size: 10px; font-weight: normal; border: 1px solid;
            border-radius: 10px; padding: 1px 6px; white-space: nowrap;
        }
        .screenshot-btn {
            background-color: #2a2a2a; color: #E5C07B; border: 1px solid #E5C07B;
            padding: 6px 16px; border-radius: 4px; cursor: pointer; font-size: 12px;
        }
        .screenshot-btn:hover { background-color: #E5C07B; color: #1b1b1b; }
        .screenshot-btn:disabled { opacity: 0.6; cursor: default; }
"""

# 分类筛选 / 板块筛选 / 顶部汇总条联动 / 历史列时间范围 / 截图导出
MATRIX_COMMON_JS = r"""
    const rows = document.querySelectorAll('tr[data-cats]');
    const summaryItems = document.querySelectorAll('.summary-item');
    const summaryTotalNum = document.getElementById('summary-total-num');
    const typeBtns = document.querySelectorAll('.type-btn');
    const boardSelect = document.getElementById('board-select');
    let curCat = '__ALL__';
    let curBoard = '';

    function updateSummary() {
        const statusCounts = {};
        let visibleTotal = 0;
        rows.forEach(function(row) {
            if (!row.classList.contains('hidden')) {
                visibleTotal++;
                const st = row.dataset.status;
                statusCounts[st] = (statusCounts[st] || 0) + 1;
            }
        });
        summaryItems.forEach(function(item) {
            const st = item.dataset.status;
            const cnt = statusCounts[st] || 0;
            item.querySelector('.summary-count').textContent = cnt;
            item.style.display = cnt === 0 ? 'none' : '';
        });
        if (summaryTotalNum) summaryTotalNum.textContent = visibleTotal;
    }

    // 一只票可能同时属于「行业」和「概念」，data-cats / data-boards 都是用 | 分隔的多值
    function applyFilter() {
        rows.forEach(function(row) {
            const cats = (row.dataset.cats || '').split('|');
            const bds = (row.dataset.boards || '').split('|');
            const okCat = (curCat === '__ALL__') || cats.indexOf(curCat) >= 0;
            const okBoard = (!curBoard) || bds.indexOf(curBoard) >= 0;
            row.classList.toggle('hidden', !(okCat && okBoard));
        });
        typeBtns.forEach(function(b) {
            b.classList.toggle('active', b.dataset.cat === curCat);
        });
        updateSummary();
    }

    typeBtns.forEach(function(btn) {
        btn.addEventListener('click', function() {
            curCat = this.dataset.cat;
            applyFilter();
        });
    });
    if (boardSelect) {
        boardSelect.addEventListener('change', function() {
            curBoard = this.value;
            applyFilter();
        });
    }
    applyFilter();

    // ---- 时间范围切换：日期列已按从新到旧排列，保留前 N 列（T+1 模拟列不受影响）----
    const rangeBtns = document.querySelectorAll('.range-btn');
    const matrixTable = document.querySelector('.table-container table');

    function applyRangeTo(cells, n) {
        let histIdx = 0;
        cells.forEach(function(cell) {
            if (!cell.classList.contains('hist-col')) return;
            histIdx++;
            cell.classList.toggle('col-hidden', histIdx > n);
        });
    }

    function setRange(n) {
        applyRangeTo(matrixTable.querySelectorAll('thead th'), n);
        matrixTable.querySelectorAll('tbody tr').forEach(function(tr) {
            applyRangeTo(tr.querySelectorAll('td'), n);
        });
        rangeBtns.forEach(function(b) {
            b.classList.toggle('active', parseInt(b.dataset.range, 10) === n);
        });
    }

    rangeBtns.forEach(function(btn) {
        btn.addEventListener('click', function() {
            setRange(parseInt(this.dataset.range, 10));
        });
    });
    const activeRangeBtn = document.querySelector('.range-btn.active');
    if (activeRangeBtn) setRange(parseInt(activeRangeBtn.dataset.range, 10));

    // ---- 截图导出 ----
    const btnScreenshot = document.getElementById('btn-screenshot');
    if (btnScreenshot) {
        btnScreenshot.addEventListener('click', function() {
            const target = document.getElementById('capture-area');
            if (!target || typeof html2canvas === 'undefined') {
                alert('截图组件加载失败，请检查网络连接后重试');
                return;
            }
            btnScreenshot.disabled = true;
            btnScreenshot.textContent = '⏳ 正在生成截图...';
            html2canvas(target, {
                backgroundColor: '#1e1e1e',
                width: target.scrollWidth,
                height: target.scrollHeight,
                windowWidth: target.scrollWidth,
                windowHeight: target.scrollHeight,
                scale: window.devicePixelRatio > 1 ? 2 : 1,
                useCORS: true
            }).then(function(canvas) {
                const link = document.createElement('a');
                const ts = new Date().toISOString().slice(0, 19).replace(/[:T]/g, '-');
                link.download = SHOT_NAME + '_' + ts + '.png';
                link.href = canvas.toDataURL('image/png');
                link.click();
                btnScreenshot.disabled = false;
                btnScreenshot.textContent = '📷 保存截图';
            }).catch(function(err) {
                console.error('截图失败:', err);
                alert('截图失败，请稍后重试');
                btnScreenshot.disabled = false;
                btnScreenshot.textContent = '📷 保存截图';
            });
        });
    }
"""

MATRIX_IND_TOGGLE_JS = r"""
    // ---- 指标切换 ----
    const indicators = ['macd', 'rsi', 'bias', 'kdj', 'cci', 'sar', 'boll', 'bw'];
    indicators.forEach(function(ind) {
        const btn = document.getElementById('btn-' + ind);
        if (!btn) return;
        btn.addEventListener('click', function() {
            document.querySelectorAll('.ind-' + ind).forEach(function(item) {
                item.classList.toggle('hidden');
            });
            this.classList.toggle('active');
        });
    });
"""

MATRIX_SIM_TOGGLE_JS = r"""
    // ---- T+1 模拟列显隐 ----
    const btnSim = document.getElementById('btn-sim');
    if (btnSim) {
        btnSim.addEventListener('click', function() {
            const on = this.classList.toggle('active');
            document.querySelectorAll('.sim-col').forEach(function(c) {
                c.classList.toggle('col-hidden', !on);
            });
        });
    }

    // ---- 评分变化数字显隐 ----
    const btnDelta = document.getElementById('btn-delta');
    if (btnDelta) {
        btnDelta.addEventListener('click', function() {
            const on = this.classList.toggle('active');
            document.querySelectorAll('.score-delta').forEach(function(c) {
                c.classList.toggle('hidden', !on);
            });
        });
    }
"""


def _range_buttons_html(label):
    btns = ''.join(
        f'<button class="range-btn{" active" if n == SHOW_HIST_DAYS else ""}" data-range="{n}">近{n}日</button>'
        for n in sorted({5, 10, 20, 30, SHOW_HIST_DAYS})
    )
    return (f'<div class="btn-group"><span style="margin-right:10px; color:#aaa;">{label}：</span>'
            f'{btns}</div>')


# ==========================================================================
# 9. ② 近六十日八大指标交互式矩阵（原 17 号脚本，分析对象换成个股）
# ==========================================================================
def _cell_color(v, red_ge=None, green_le=None):
    if v is None or (isinstance(v, float) and np.isnan(v)):
        return "#888888"
    if red_ge is not None and v >= red_ge:
        return "#FF3333"
    if green_le is not None and v <= green_le:
        return "#00CC00"
    return "#FFFFFF"


def build_matrix_17(rows, out_path):
    dates_set = set()
    for r in rows:
        dates_set.update(r['ind'].index)
    sorted_dates = sorted(dates_set, reverse=True)[:SHOW_HIST_DAYS]
    date_strs = [d.strftime('%Y-%m-%d') for d in sorted_dates]

    html = f"""<!DOCTYPE html>
<html>
<head>
    <meta charset="utf-8">
    <meta name="viewport" content="width=device-width, initial-scale=1">
    <title>个股近{SHOW_HIST_DAYS}日八大指标交互式矩阵</title>
    <style>{MATRIX_CSS}</style>
</head>
<body>
    <h2>个股近{SHOW_HIST_DAYS}日核心指标矩阵（MACD, RSI, BIAS, KDJ, CCI, SAR, BOLL, BW%）</h2>
    <div class="subtitle">
        数据源：akshare 前复权日线；指标口径与 18 号波段内核完全一致。
        分类按板块类别（行业 / 概念）划分，一只票横跨多个板块时几个分类都能筛到它。
    </div>

    {_summary_bar_html(rows)}

    {_filter_bar_html(rows)}

    {_range_buttons_html('时间范围')}

    <!-- 指标切换按钮 + 截图导出按钮（同一行，截图按钮靠右） -->
    <div class="btn-group indicator-row">
        <div class="indicator-row-left">
            <span style="margin-right:10px; color:#aaa;">指标：</span>
            <button id="btn-macd" class="active">MACD</button>
            <button id="btn-rsi"  class="active">RSI</button>
            <button id="btn-bias" class="active">BIAS</button>
            <button id="btn-kdj"  class="active">KDJ</button>
            <button id="btn-cci"  class="active">CCI</button>
            <button id="btn-sar"  class="active">SAR</button>
            <button id="btn-boll" class="active">BOLL</button>
            <button id="btn-bw"   class="active">BW%</button>
        </div>
        <button id="btn-screenshot" class="screenshot-btn">📷 保存截图</button>
    </div>

    <div class="table-container" id="capture-area">
        <table>
            <thead>
                <tr>
                    <th class="index-name-th">股票 \\ 日期</th>
"""
    for ds in date_strs:
        html += f'<th class="hist-col">{ds}</th>'
    html += "</tr></thead><tbody>"

    for r in rows:
        df_ind = r['ind']
        slug = STATUS_SLUG_MAP.get(r['heat'].get('status', '⚪ 无数据'), 'nodata')
        html += (f'<tr data-cats="{_esc("|".join(r["cats"]))}" '
                 f'data-boards="{_esc("|".join(r["boards"]))}" data-status="{slug}">')
        html += _row_head_cell(r)

        for d in sorted_dates:
            if d not in df_ind.index:
                html += '<td class="hist-col">-</td>'
                continue
            row = df_ind.loc[d]
            macd = float(row.get('macd_hist', np.nan))
            rsi = float(row.get('rsi', np.nan))
            bias = float(row.get('bias', np.nan))
            kdj = float(row.get('kdj', np.nan))
            cci = float(row.get('cci', np.nan))
            sar = str(row.get('sar_stat', '中性'))
            boll = float(row.get('boll_pctb', np.nan))
            bw_pct = float(row.get('boll_bw_pct', np.nan))

            c_macd = "#888888" if np.isnan(macd) else ("#FF3333" if macd >= 0 else "#00CC00")
            c_rsi = _cell_color(rsi, 70, 30)
            c_bias = _cell_color(bias, 4, -4)
            c_kdj = _cell_color(kdj, 90, 10)
            c_cci = _cell_color(cci, 100, -100)
            c_sar = "#c0392b" if sar == "看涨" else "#27ae60"
            c_boll = _cell_color(boll, 0.9, 0.1)
            c_bw = _cell_color(bw_pct, 0.8, 0.2)

            # 判断全红 / 全绿（RSI, BIAS, KDJ, CCI, SAR, BOLL 六项共振）
            all_red = (c_rsi == '#FF3333' and c_bias == '#FF3333' and c_kdj == '#FF3333'
                       and c_cci == '#FF3333' and c_sar == '#c0392b' and c_boll == '#FF3333')
            all_green = (c_rsi == '#00CC00' and c_bias == '#00CC00' and c_kdj == '#00CC00'
                         and c_cci == '#00CC00' and c_sar == '#27ae60' and c_boll == '#00CC00')
            bg_style = ""
            if all_red:
                bg_style = ' style="background-color: rgba(255,0,0,0.2);"'
            elif all_green:
                bg_style = ' style="background-color: rgba(0,255,0,0.2);"'

            cell_html = f"""
            <div class="indicators-grid">
                <div class="indicator-item ind-macd"><span class="ind-label">MACD</span><span class="ind-val" style="color:{c_macd};">{_fmt(macd, '+.3f')}</span></div>
                <div class="indicator-item ind-rsi"><span class="ind-label">RSI</span><span class="ind-val" style="color:{c_rsi};">{_fmt(rsi, '.1f')}</span></div>
                <div class="indicator-item ind-bias"><span class="ind-label">BIAS</span><span class="ind-val" style="color:{c_bias};">{_fmt(bias, '+.2f', '%')}</span></div>
                <div class="indicator-item ind-kdj"><span class="ind-label">KDJ</span><span class="ind-val" style="color:{c_kdj};">{_fmt(kdj, '.1f')}</span></div>
                <div class="indicator-item ind-cci"><span class="ind-label">CCI</span><span class="ind-val" style="color:{c_cci};">{_fmt(cci, '.1f')}</span></div>
                <div class="indicator-item ind-sar"><span class="ind-label">SAR</span><span class="ind-val" style="color:{c_sar};">{sar}</span></div>
                <div class="indicator-item ind-boll"><span class="ind-label">BOLL</span><span class="ind-val" style="color:{c_boll};">{_fmt(boll, '.3f')}</span></div>
                <div class="indicator-item ind-bw"><span class="ind-label">BW%</span><span class="ind-val" style="color:{c_bw};">{_fmt(np.nan if np.isnan(bw_pct) else bw_pct * 100, '.0f', '%')}</span></div>
            </div>
            """
            html += f'<td class="hist-col"{bg_style}>{cell_html}</td>'
        html += "</tr>"

    html += """
        </tbody>
    </table>
</div>

<script src="https://cdnjs.cloudflare.com/ajax/libs/html2canvas/1.4.1/html2canvas.min.js"></script>
<script>
var SHOT_NAME = '个股指标矩阵';
document.addEventListener('DOMContentLoaded', function() {
""" + MATRIX_COMMON_JS + MATRIX_IND_TOGGLE_JS + """
});
</script>
</body>
</html>
"""
    with open(out_path, 'w', encoding='utf-8') as f:
        f.write(html)
    print(f"✅ 成功生成个股指标矩阵: {out_path}")


# ==========================================================================
# 10. ④ 过热评分矩阵 + T+1 情景模拟（原 19 号脚本，分析对象换成个股）
# ==========================================================================
def render_badge(status, color, score, delta=None, tip='', extra_class=''):
    """统一的评分徽章渲染：🟡 偏热 · 69分 （+6）"""
    bg = _hex_to_rgba(color, 0.14)
    if score is None or (isinstance(score, float) and np.isnan(score)):
        score_txt = '--'
    else:
        score_txt = f'{score:.0f}分'
    delta_html = ''
    if delta is not None and not (isinstance(delta, float) and np.isnan(delta)):
        if abs(delta) < 0.5:
            d_color, d_sign = '#888888', '±0'
        elif delta > 0:
            d_color, d_sign = '#FF6B6B', f'+{delta:.0f}'
        else:
            d_color, d_sign = '#4ADE80', f'{delta:.0f}'
        delta_html = f'<span class="score-delta" style="color:{d_color};">{d_sign}</span>'
    tip_attr = f' title="{tip}"' if tip else ''
    return (
        f'<div class="score-badge {extra_class}" '
        f'style="color:{color};border-color:{color};background:{bg};"{tip_attr}>'
        f'<span class="score-status">{status}</span>'
        f'<span class="score-num">{score_txt}</span>{delta_html}</div>'
    )


def build_matrix_19(rows, out_path):
    C = core()
    dates_set = set()
    for r in rows:
        dates_set.update(r['ind'].index)
    sorted_dates = sorted(dates_set, reverse=True)[:SHOW_HIST_DAYS]
    date_strs = [d.strftime('%Y-%m-%d') for d in sorted_dates]
    sim_col_count = len(SIM_RETURNS)

    html = f"""<!DOCTYPE html>
<html>
<head>
    <meta charset="utf-8">
    <meta name="viewport" content="width=device-width, initial-scale=1">
    <title>个股过热评分矩阵 + T+1 情景模拟</title>
    <style>{MATRIX_CSS}</style>
</head>
<body>
    <h2>个股过热评分矩阵（近{SHOW_HIST_DAYS}日） + 下一交易日涨跌情景模拟</h2>
    <div class="subtitle">
        单元格 = 当日八大指标算出的过热评分与状态；鼠标悬停可看八大指标明细。
        左侧 T+1 列 = 假设下一交易日涨跌该幅度 → 推算收盘价 → 重算指标 → 重算评分，括号内为相对今日评分的变化。
    </div>

    {_summary_bar_html(rows)}

    {_filter_bar_html(rows)}

    {_range_buttons_html('历史范围')}

    <div class="btn-group tool-row">
        <div class="tool-row-left">
            <span style="margin-right:10px; color:#aaa;">显示：</span>
            <button id="btn-sim" class="active">T+1 模拟列</button>
            <button id="btn-delta" class="active">评分变化</button>
        </div>
        <button id="btn-screenshot" class="screenshot-btn">📷 保存截图</button>
    </div>

    <div class="table-container" id="capture-area">
        <table>
            <thead>
                <tr>
                    <th class="index-name-th">股票 \\ 日期</th>
"""
    for i, pct in enumerate(SIM_RETURNS):
        cls = 'sim-col'
        if i == 0:
            cls += ' sim-first'
        if i == sim_col_count - 1:
            cls += ' sim-last'
        sign = '+' if pct > 0 else ''
        html += (f'<th class="{cls}">'
                 f'<span class="sim-head-main">T+1 {sign}{pct:.0f}%</span>'
                 f'<span class="sim-head-sub">模拟</span></th>')
    for ds in date_strs:
        html += f'<th class="hist-col">{ds}</th>'
    html += "</tr></thead><tbody>"

    for r in rows:
        df_ind = r['ind']
        slug = STATUS_SLUG_MAP.get(r['heat'].get('status', '⚪ 无数据'), 'nodata')
        html += (f'<tr data-cats="{_esc("|".join(r["cats"]))}" '
                 f'data-boards="{_esc("|".join(r["boards"]))}" data-status="{slug}">')
        html += _row_head_cell(r)

        sims = r.get('sims') or {}
        for i, pct in enumerate(SIM_RETURNS):
            cls = 'sim-col'
            if i == 0:
                cls += ' sim-first'
            if i == sim_col_count - 1:
                cls += ' sim-last'
            s = sims.get(pct)
            if not s:
                html += f'<td class="{cls}">-</td>'
                continue
            badge = render_badge(s['status'], s['color'], s['score'],
                                 delta=s['delta'], tip=s['tip'], extra_class='sim-badge')
            html += f'<td class="{cls}">{badge}</td>'

        for d in sorted_dates:
            if d in df_ind.index:
                row = df_ind.loc[d]
                sc = row.get('score', np.nan)
                sc_val = None if pd.isna(sc) else float(sc)
                st, mn, ac, cl = C.get_status_action(sc_val)
                badge = render_badge(st, cl, sc_val, delta=None, tip=_indicator_tip(row))
                html += f'<td class="hist-col">{badge}</td>'
            else:
                html += '<td class="hist-col">-</td>'
        html += "</tr>"

    html += """
        </tbody>
    </table>
</div>

<script src="https://cdnjs.cloudflare.com/ajax/libs/html2canvas/1.4.1/html2canvas.min.js"></script>
<script>
var SHOT_NAME = '个股评分矩阵';
document.addEventListener('DOMContentLoaded', function() {
""" + MATRIX_COMMON_JS + MATRIX_SIM_TOGGLE_JS + """
});
</script>
</body>
</html>
"""
    with open(out_path, 'w', encoding='utf-8') as f:
        f.write(html)
    print(f"✅ 成功生成个股评分矩阵: {out_path}")


# ==========================================================================
# 11. ③ ZigZag 波浪 + 波段信号量化看板（原 18 号脚本的主程序，改喂个股）
# ==========================================================================
def build_zigzag_18(rows, out_path):
    C = core()
    items = [r for r in rows if r.get('zz')]
    if not items:
        print("⚠️ 没有可用于波段看板的标的，已跳过")
        return

    # 顶部「板块」下拉：全部 → 具体板块（比如 A股宽基/港股宽基/美股宽基/外盘宽基）。
    # BROAD_INDEX_ONLY 模式下类别永远只有"宽基"一种，类别级选项和"全部"完全重复，
    # 所以这里只保留"全部" + 具体板块两级，不再单独列类别。
    cats, cat_boards = [], {}
    for r in items:
        for b, c in r['bpairs']:
            if c not in cats:
                cats.append(c)
            if b:
                cat_boards.setdefault(c, set()).add(b)

    def keys_of(pred):
        picked = [r for r in items if pred(r)]
        picked.sort(key=lambda r: r['heat'].get('score', 0), reverse=True)
        return [r['zz']['key'] for r in picked]

    all_types = ['全部']
    type_funds = {'全部': keys_of(lambda r: True)}
    for c in cats:
        for b in sorted(cat_boards.get(c, [])):
            label = b if len(cats) <= 1 else f"{c}·{b}"
            type_funds[label] = keys_of(lambda r, b=b: b in r['boards'])
            all_types.append(label)

    nav_charts_html, score_charts_html = [], []
    fund_meta, fund_fixed, fund_pct, fund_signals, fund_scores = {}, {}, {}, {}, {}
    for r in sorted(items, key=lambda r: r['heat'].get('score', 0), reverse=True):
        z = r['zz']
        k = z['key']
        nav_charts_html.append(z['nav_html'])
        score_charts_html.append(z['score_html'])
        fund_meta[k] = z['meta']
        fund_fixed[k] = z['fixed']
        fund_pct[k] = z['pct']
        fund_signals[k] = z['signal_latest']
        fund_scores[k] = z['score_history']

    indicator_cells = '\n'.join(
        f'        <div class="ind-cell"><span class="ind-label">{label}</span>'
        f'<span class="ind-value" id="ind-{key}">--</span></div>'
        for key, label in C.INDICATOR_LABELS
    )
    pct_strip = '\n'.join(
        f'        <div class="pct-pill"><span class="pct-label">{C.WINDOW_CONFIG[k]["label"]}</span>'
        f'<span class="pct-value" id="pct-value-{k}">--</span></div>' for k in C.WINDOW_ORDER
    )

    # 内核模板是给基金写的，这里只改文案，结构与交互一字不动
    html = C.HTML_TEMPLATE
    html = html.replace('基金波段信号量化看板', '个股波段信号量化看板')
    html = html.replace('>基金</label>', '>股票</label>')
    html = html.replace('净值走势', '价格走势')

    page_js = C.PAGE_JS.replace(
        "'） 最新净值 ' + meta.last_nav.toFixed(4)",
        "'） 最新价 ' + meta.last_nav.toFixed(2)")

    html = html.replace('__NAV_CHARTS__', '\n'.join(nav_charts_html))
    html = html.replace('__INDICATOR_CELLS__', indicator_cells)
    html = html.replace('__PCT_STRIP__', pct_strip)
    html = html.replace('__SCORE_CHARTS__', '\n'.join(score_charts_html))
    html = html.replace('__GEN_TIME__', pd.Timestamp.now().strftime('%Y-%m-%d %H:%M'))
    html = html.replace('__ZIGZAG_PCT__', f'{C.ZIGZAG_THRESHOLD * 100:.0f}')
    html = html.replace('__CSS__', C.PAGE_CSS)
    html = html.replace('__JS_BLOCK__', page_js)
    html = html.replace('__ALL_TYPES_JSON__', json.dumps(all_types, ensure_ascii=False))
    html = html.replace('__TYPE_FUNDS_JSON__', json.dumps(type_funds, ensure_ascii=False))
    html = html.replace('__FUND_META_JSON__', json.dumps(fund_meta, ensure_ascii=False))
    html = html.replace('__FUND_FIXED_JSON__', json.dumps(fund_fixed, ensure_ascii=False))
    html = html.replace('__FUND_PCT_JSON__', json.dumps(fund_pct, ensure_ascii=False))
    html = html.replace('__FUND_SIGNALS_JSON__', json.dumps(fund_signals, ensure_ascii=False))
    html = html.replace('__FUND_SCORE_HISTORIES_JSON__', json.dumps(fund_scores, ensure_ascii=False))

    with open(out_path, 'w', encoding='utf-8') as f:
        f.write(html)
    print(f"✅ 成功生成个股波段信号看板: {out_path}（收录 {len(items)} 只标的）")


# ==========================================================================
# 12. 抓取(线程) → 分析(进程) 流水线
# ==========================================================================
def run_pipeline(tasks, cfg, t_start):
    P = pm()
    P.FETCH_RETRY = FETCH_RETRY
    total = len(tasks)
    results, state = [], {'done': 0}

    def collect(res):
        state['done'] += 1
        if res is None:
            return
        if res.get('err'):
            print(res['err'])
        results.append(res)
        if state['done'] % 10 == 0 or state['done'] == total:
            print(f"    —— 进度 {state['done']}/{total}，已用时 {time.time() - t_start:.1f}s")

    pool_exec = None
    try:
        if USE_PROCESS_POOL and RENDER_WORKERS > 1:
            pool_exec = ProcessPoolExecutor(max_workers=RENDER_WORKERS)
    except Exception as e:
        print(f"⚠️ 进程池创建失败，退回单进程: {e}")
        pool_exec = None

    analyze_futures = []
    with ThreadPoolExecutor(max_workers=FETCH_WORKERS) as tp:
        fetch_futures = [tp.submit(P.fetch_task, t) for t in tasks]
        for fut in as_completed(fetch_futures):
            task, df = fut.result()
            if df is None or df.empty:
                state['done'] += 1
                print(f"❌ {task['name']} ({task['code']}) 数据获取失败，跳过。")
                continue
            print(f">>> 已获取: [{task['region']}]【{task.get('sector_display') or task['sector']}】 "
                  f"{task['name']} ({task['code']})")
            if pool_exec is not None:
                analyze_futures.append(pool_exec.submit(analyze_one, (task, df, cfg)))
            else:
                collect(analyze_one((task, df, cfg)))

    if pool_exec is not None:
        for fut in as_completed(analyze_futures):
            try:
                collect(fut.result())
            except Exception as e:
                state['done'] += 1
                print(f"❌ 分析子进程异常: {e}")
        pool_exec.shutdown(wait=True)

    return results


# ==========================================================================
# 13. 主运行流
# ==========================================================================
def main():
    t_start = time.time()
    P = pm()
    os.makedirs(CHART_DIR, exist_ok=True)
    os.makedirs(DASHBOARD_DIR, exist_ok=True)

    # ---------- 1) 读板块池 → 勾选板块与输出内容 ----------
    if BROAD_INDEX_ONLY:
        pool = pd.DataFrame()
        print(">>> 开源版：只分析宽基指数，跳过板块个股股票池")
    else:
        print(">>> 正在读取板块股票池...")
        pool = P.load_board_pool(BOARD_CSV)
        if pool.empty:
            print(f"⚠️ {BOARD_CSV} 为空或不可解析，只能分析国外标的 / 宽基指数")

    choice = choose_options(pool)
    boards, top_n = choice['boards'], choice['top_n']
    if boards:
        shown = '、'.join(boards[:12]) + ('…' if len(boards) > 12 else '')
        print(f">>> 已选【{choice.get('category')}】{len(boards)} 个板块：{shown}")
        print(f">>> 每板块取前 {top_n if top_n else '全部'} 只（按基金持有比例(%) 从高到低）")
    else:
        print("⚠️ 未选到任何板块，本次只分析国外标的与宽基指数")
    picked = [n for n, on in (('形态图', choice['charts']), ('板块Excel', choice['excel']),
                              ('指标矩阵', choice['m17']), ('波段看板', choice['m18']),
                              ('评分矩阵', choice['m19'])) if on]
    print(f">>> 输出内容：{'、'.join(picked) if picked else '（无）'}")

    # ---------- 2) 组装标的（按代码去重：一只票只抓一次数、只画一张图）----------
    tasks = build_tasks(pool, choice)
    if not tasks:
        print("❌ 没有可分析的标的，流程结束。")
        return

    multi = sum(1 for t in tasks if len(t['boards']) > 1)
    pair_cnt = sum(len(t['boards']) for t in tasks)
    print(f">>> 共 {len(tasks)} 个标的（去重后）｜其中 {multi} 只横跨多个板块，"
          f"板块-成分对 {pair_cnt} 条｜抓取并发 {FETCH_WORKERS} 线程，"
          f"分析并发 {RENDER_WORKERS if USE_PROCESS_POOL else 1} 进程")

    cfg = dict(
        charts=choice['charts'], excel=choice['excel'],
        m17=choice['m17'], m18=choice['m18'], m19=choice['m19'],
        chart_dir=CHART_DIR,
        signal_days=P.SIGNAL_RECENT_DAYS,
        plot_bars=P.PLOT_RECENT_BARS,
        plotly_js=P.PLOTLY_JS_MODE,
        max_bars=ANALYSIS_MAX_BARS,
        hist_days=SHOW_HIST_DAYS,
        sim_returns=list(SIM_RETURNS),
    )

    # ---------- 3) 抓取 + 分析 ----------
    results = run_pipeline(tasks, cfg, t_start)
    print(f"\n>>> 取数与计算阶段完成，用时 {time.time() - t_start:.1f}s")

    # ---------- 4) 出报表 ----------
    if choice['excel']:
        write_pattern_excel([r['signal_df'] for r in results if r.get('signal_df') is not None],
                            CHART_DIR)

    rows = collect_rows(results)
    if not rows:
        print("⚠️ 没有可用于矩阵/看板的指标数据（可能全部标的数据量不足）")
    else:
        if choice['m17']:
            build_matrix_17(rows, os.path.join(DASHBOARD_DIR, HTML_MATRIX))
        if choice['m19']:
            build_matrix_19(rows, os.path.join(DASHBOARD_DIR, HTML_SCORE))
        if choice['m18']:
            build_zigzag_18(rows, os.path.join(DASHBOARD_DIR, HTML_ZIGZAG))

    print(f"\n🎉 全部完成，总用时 {time.time() - t_start:.1f}s")
    print(f"   K线形态图与 Excel：{os.path.abspath(CHART_DIR)}")
    print(f"   看板 HTML：{os.path.abspath(DASHBOARD_DIR)}")


if __name__ == '__main__':
    mp.freeze_support()          # Windows 打包 / spawn 必需
    main()