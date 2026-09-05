# -*- coding: utf-8 -*-
"""
自动化批处理股票池 (读取CSV)
严格识别15种看跌与7种看涨形态，包含指定样式的布林带，量价分析
新增：智能识别并用半透明蓝色矩形绘制 向上/向下跳空缺口
新增：支持读取并展示所属“板块”信息
新增：支持美股、港股、全球宽基指数数据的获取
新增：汇总所有个股与指数信号，按“板块”输出每日看多/看空/缺口情况的 Excel 表格

【本次改造】
1. 国内股票池改读 ztjj_board_stocks.csv（板块名称 / 板块类别 / 股票代码 / 股票名称 / 基金持有比例）
   板块类别分“行业”与“概念”两类，运行前弹出勾选对话框（无图形界面时自动退回命令行选择）
2. 默认按“概念”，并预勾选 DEFAULT_BOARDS 里的板块（其中有色金属/电力/白酒/银行属“行业”，
   切到“行业”或“全部”时同样会被预勾选）
3. 一只股票可能同时属于多个板块 —— 按股票代码去重，数据只抓一次、图只画一张，
   信号在汇总阶段再复制到它所属的每一个板块（不会重复抓数、不会重复渲染）
4. 国外标的（target_stocks.csv 中 global=国外 的行）保持原样不变
5. 汇总 Excel 新增“板块类别”列，Sector_heatmap.py 据此增加 行业/概念 切换
"""

import os
import time
import pandas as pd
import numpy as np
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import warnings
import openpyxl
from openpyxl.styles import PatternFill
from concurrent.futures import ThreadPoolExecutor, ProcessPoolExecutor, as_completed
import multiprocessing as mp
warnings.filterwarnings('ignore')

# akshare 导入很慢（数秒），且只有抓数据时用得到。
# 延迟到函数内部导入，绘图子进程就不必再加载它，能省下大量进程启动时间。

# ----------------------------- 0.0 性能与并发配置 -----------------------------
# 抓数据是网络IO密集 → 用线程；画图是CPU密集(受GIL限制) → 用进程
FETCH_WORKERS = 8                                   # 数据抓取并发线程数（太大易被限流）
FETCH_RETRY = 2                                     # 抓取失败的重试次数（并发下接口偶发抽风）
RENDER_WORKERS = max(1, (os.cpu_count() or 4) - 1)  # 绘图并发进程数，默认留一个核给系统
USE_PROCESS_POOL = True                             # 出问题时置 False 退回单进程绘图

# plotly.js 引入方式：
#   'directory' → 输出目录放一份 plotly.min.js，各 HTML 引用它（推荐：快、体积小、可离线）
#   'cdn'       → 从网上加载，文件最小，但打开图表时需要联网
#   True        → 每个 HTML 内嵌一份 3MB 的 js（原行为，最慢）
PLOTLY_JS_MODE = "directory"

# 只绘制最近 N 根K线（0 = 全部历史）。指标与形态仍按全量历史计算，判定结果不受影响。
# A股老股票动辄 5000+ 根K线，画全量会让 HTML 又大又卡；建议 500~1000。
PLOT_RECENT_BARS = 750

# 汇总表只保留最近 N 个交易日的信号（0 = 全部）。板块报表默认只看 30 天，
# 留 120 天足够用，能大幅缩小 Excel 体积与写入时间。
# IC 检验需要长时序：120 天切成 3 日窗口只剩 ~40 个独立样本，10 日只剩 ~11 个。
# 做因子研究时设 500+；日常只出报表可调回 120 以缩小 Excel。
SIGNAL_RECENT_DAYS = 500

# 多空计数口径：
#   'v1' = 看涨+向上缺口+量价看多 / 看跌+向下缺口+量价看空（原始口径）
#   'v2' = 向上缺口+量价看多 / 向下缺口+量价看空（剔除未通过稳定性检验的形态项）
SIGNAL_SCHEME = 'v1'

# ----------------------------- 0.A 股票池 / 板块配置 -----------------------------
BOARD_CSV = "ztjj_board_stocks.csv"     # 国内股票池：板块名称/板块类别/股票代码/股票名称/基金持有比例(%)
LEGACY_CSV = "target_stocks.csv"        # 老股票池：仍用于取“国外”标的（保持不变）

# 运行前是否弹出勾选对话框；False 则直接用下面的默认值静默运行（适合定时任务）
USE_SELECTION_UI = True

# 默认板块类别：'概念' / '行业' / '全部'
DEFAULT_BOARD_CATEGORY = "概念"

# 默认勾选的板块（跨类别写在一起：切到哪个类别就自动勾选属于该类别的那几个）
DEFAULT_BOARDS = [
    "CPO", "存储芯片", "第三代半导体", "有色金属", "稀土永磁", "CRO","PCB",
    "创新药", "电力", "白酒", "黄金股", "锂矿", "银行",
]

# 每个板块最多取多少只成分股（按“基金持有比例(%)”从高到低，0 = 全部）
# ztjj 里一个概念动辄 100~700 只票，全取会让抓数与绘图时间爆炸，默认取机构重仓的前 30 只
DEFAULT_TOP_N_PER_BOARD = 30

# 是否为每只票生成 K 线 HTML。选中多个概念时标的数可能上百，只想要板块统计表时可置 False
GENERATE_CHARTS = True

INCLUDE_OVERSEAS = True          # 是否并入 target_stocks.csv 的国外标的（默认并入，口径不变）
INCLUDE_LEGACY_DOMESTIC = False  # 是否同时并入 target_stocks.csv 的国内标的（默认不并，避免与 ztjj 重复）

# 国外板块名 → ztjj 板块名 的对齐表：只影响“全球合并”时能不能跟 A 股同名板块合到一行，
# 不改变国外标的本身。设为 {} 即完全保持旧口径
OVERSEAS_BOARD_ALIAS = {
    "有色": "有色金属",
    "稀土": "稀土永磁",
    "医疗": "医疗服务",
    "芯片设计": "半导体",
    "半导体制造封测": "半导体",
    "半导体设备材料": "半导体",
    "算力服务器": "算力",
}
# 国外板块在 ztjj 里查不到类别时的兜底归类
OVERSEAS_FALLBACK_CATEGORY = "概念"

# ----------------------------- 0. 全局宽基指数配置 -----------------------------
# 结构：[名称, 接口类型, 接口代码, 归属(国内/国外), 宽基分组]
# 归属与 CSV 中新增的 global 列口径保持一致，用于报表按 国内/国外/全球 拆分
BROAD_INDICES = {
    # --- A股指数 ---
    "sh000001": ["上证指数", "zh", "sh000001", "国内", "A股宽基"],
    "sh000300": ["沪深300", "zh", "sh000300", "国内", "A股宽基"],
    "sz399001": ["深证成指", "zh", "sz399001", "国内", "A股宽基"],
    "sz399006": ["创业板指", "zh", "sz399006", "国内", "A股宽基"],
    "sh000016": ["上证50", "zh", "sh000016", "国内", "A股宽基"],
    "sh000905": ["中证500", "zh", "sh000905", "国内", "A股宽基"],
    "sh000688": ["科创50", "zh", "sh000688", "国内", "A股宽基"],
    "bj899050": ["北证50", "zh", "bj899050", "国内", "A股宽基"],
    "sh000010": ["上证收益", "zh", "sh000010", "国内", "A股宽基"],

    # --- 美股指数 ---
    ".IXIC": ["纳斯达克", "us", ".IXIC", "国外", "美股宽基"],
    ".INX": ["标普500", "us", ".INX", "国外", "美股宽基"],
    ".DJI": ["道琼斯", "us", ".DJI", "国外", "美股宽基"],

    # --- 港股指数 ---
    "hkHSI": ["恒生指数", "hk", "HSI", "国外", "港股宽基"],
    "hkHSTECH": ["恒生科技指数", "hk", "HSTECH", "国外", "港股宽基"],

    # --- 环球外盘指数 ---
    "jpN225": ["日经225", "global", "日经225指数", "国外", "外盘宽基"],
    "krKOSPI": ["韩国综合指数", "global", "首尔综合指数", "国外", "外盘宽基"]
}

# ----------------------------- 0.1 量价关系纳入多空统计的口径 -----------------------------
# 原口径：利好 = 看涨形态 + 向上缺口；利空 = 看跌形态 + 向下缺口
# 新口径：额外计入量价关系（只采信“放量”这类有资金验证的强信号，
#         缩量上涨/缩量下跌属于中性偏弱，不计入，避免噪声灌水）
# 如需把“缩量下跌 (抛压减轻)”也视作利好，把它加进 VP_BULLISH_KEYS 即可
VP_BULLISH_KEYS = ["放量上涨"]      # 量价看多
VP_BEARISH_KEYS = ["放量下跌"]      # 量价看空

# ----------------------------- 0.2 涨跌停统计口径 -----------------------------
# 涨跌停只对 A 股有意义（美股/港股/指数无涨跌幅限制，一律记 0）
#   主板/中小板 10%｜创业板(sz30)/科创板(sh688) 20%｜北交所(bj) 30%｜主板 ST 5%
# 判定用【涨幅比例】而非重算限价：数据取的是前复权价，前复权后价格不再是当日真实
# 成交价，用 round(prev*1.1, 2) 反推限价会大面积误判，比例法则不受复权缩放影响。
LIMIT_TOL_DOWN = 0.005   # 允许的向下容差：限价要按分钱取整，实际涨幅常略小于名义值
LIMIT_TOL_UP = 0.012     # 允许的向上容差：超出即视为无涨跌停限制(新股首日/异常数据)，不计
LIMIT_SEAL_ONLY = True   # True = 只统计收盘封住的板（收盘价=最高价），炸板不计
LIMIT_SEAL_TOL = 0.001   # "收盘=最高"的相对容差

def limit_rate(code: str, name: str = "") -> float:
    """返回该标的的单日涨跌幅限制比例；0 表示无涨跌停制度，不参与统计"""
    c = (code or "").lower()
    n = (name or "").upper().replace(" ", "")
    if c in BROAD_INDICES:
        return 0.0
    if not c.startswith(("sh", "sz", "bj")):
        return 0.0                      # 美股 / 港股
    if c.startswith("bj"):
        return 0.30
    if c.startswith("sh688") or c.startswith("sz30"):
        return 0.20                     # 科创板 / 创业板（含其 ST，仍为 20%）
    if "ST" in n:
        return 0.05                     # 主板 ST / *ST
    return 0.10

# ----------------------------- 1. 数据获取与指标计算 -----------------------------
def calculate_sar(close_series, af_start=0.02, af_step=0.02, af_max=0.2):
    """
    标准抛物线SAR算法（Wilder's Parabolic SAR），与5号(基金策略)、6号(大盘指数)
    脚本使用完全相同的算法，保证SAR判断口径统一
    """
    close = close_series.values
    sar = np.zeros(len(close))
    trend = np.ones(len(close))
    ep = close[0]
    af = af_start
    sar[0] = close[0]
    for i in range(1, len(close)):
        sar[i] = sar[i-1] + af * (ep - sar[i-1])
        if trend[i-1] == 1:
            if close[i] < sar[i]:
                trend[i], sar[i], af, ep = -1, ep, af_start, close[i]
            else:
                trend[i] = 1
                if close[i] > ep: ep, af = close[i], min(af + af_step, af_max)
        else:
            if close[i] > sar[i]:
                trend[i], sar[i], af, ep = 1, ep, af_start, close[i]
            else:
                trend[i] = -1
                if close[i] < ep: ep, af = close[i], min(af + af_step, af_max)
    return sar, trend

def get_stock_data(code: str, name: str = "") -> pd.DataFrame:
    """自动路由数据接口：整合个股与各类宽基指数

    name 只用于识别 ST（主板 ST 涨跌停幅度为 5%），不影响取数逻辑"""
    import akshare as ak   # 延迟导入：只有抓数据的线程需要，绘图子进程不加载
    try:
        if code in BROAD_INDICES:
            name, api_type, symbol = BROAD_INDICES[code][:3]
            if api_type == "zh":
                df = ak.stock_zh_index_daily(symbol=symbol)
            elif api_type == "us":
                df = ak.index_us_stock_sina(symbol=symbol)
            elif api_type == "hk":
                df = ak.stock_hk_index_daily_sina(symbol=symbol)
            elif api_type == "global":
                # 使用修正后的稳定接口
                df = ak.index_global_hist_sina(symbol=symbol)
            else:
                return pd.DataFrame()
        elif code.startswith(("sh", "sz", "bj")):
            # 个股数据接口[cite: 3]
            df = ak.stock_zh_a_daily(symbol=code, adjust="qfq")
        else:
            df = ak.stock_us_daily(symbol=code, adjust="")
            
        if df is None or df.empty:
            return pd.DataFrame()
            
        # 统一标准化列名映射 (兼容大小写及不同接口的返回字段)
        col_map = {}
        for c in df.columns:
            cl = c.lower()
            if cl == 'date': col_map[c] = 'Date'
            elif cl == 'open': col_map[c] = 'Open'
            elif cl == 'high': col_map[c] = 'High'
            elif cl == 'low': col_map[c] = 'Low'
            elif cl == 'close': col_map[c] = 'Close'
            elif cl == 'volume': col_map[c] = 'Volume'
        
        df.rename(columns=col_map, inplace=True)
        
        if 'Date' not in df.columns or 'Close' not in df.columns:
            return pd.DataFrame()
            
        # 外盘指数可能没有成交量，补充默认值防崩溃
        if 'Volume' not in df.columns:
            df['Volume'] = 0
    
        df['Date'] = pd.to_datetime(df['Date'])
        df.set_index('Date', inplace=True)
        df.sort_index(inplace=True)
        
        # 确保数据格式为数值型[cite: 3]
        for col in ['Open', 'High', 'Low', 'Close', 'Volume']:
            df[col] = pd.to_numeric(df[col], errors='coerce')
        df.dropna(subset=['Close'], inplace=True)
        
        # 计算均线[cite: 3]
        df['MA5'] = df['Close'].rolling(window=5).mean()
        df['MA10'] = df['Close'].rolling(window=10).mean()
        df['MA20'] = df['Close'].rolling(window=20).mean()
        
        # 计算布林带[cite: 3]
        df['STD20'] = df['Close'].rolling(window=20).std()
        df['boll_upper'] = df['MA20'] + 2 * df['STD20']
        df['boll_lower'] = df['MA20'] - 2 * df['STD20']

        # 量价分析逻辑计算[cite: 3]
        prev_close = df['Close'].shift(1)
        prev_vol = df['Volume'].shift(1)
        vol_ma5 = df['Volume'].rolling(window=5).mean()
        
        conditions = [
            (df['Close'] > prev_close) & (df['Volume'] > prev_vol) & (df['Volume'] > vol_ma5),
            (df['Close'] > prev_close) & (df['Volume'] < prev_vol),
            (df['Close'] < prev_close) & (df['Volume'] > prev_vol) & (df['Volume'] > vol_ma5),
            (df['Close'] < prev_close) & (df['Volume'] < prev_vol),
            (df['Close'] == prev_close)
        ]
        choices = ['放量上涨 (资金积极)', '缩量上涨 (追高不足)', '放量下跌 (抛压沉重)', '缩量下跌 (抛压减轻)', '平盘震荡']
        df['VP_Analysis'] = np.select(conditions, choices, default='温和震荡')

        # 量价关系信号旗标（供板块多空统计使用）
        df['VP_Bull'] = df['VP_Analysis'].str.contains('|'.join(VP_BULLISH_KEYS), na=False).astype(int)
        df['VP_Bear'] = df['VP_Analysis'].str.contains('|'.join(VP_BEARISH_KEYS), na=False).astype(int)
        # 指数类数据若无成交量则量价关系无意义，直接置零
        if (df['Volume'].fillna(0) <= 0).all():
            df['VP_Bull'] = 0
            df['VP_Bear'] = 0
        
        # --- 缺口计算逻辑 ---[cite: 3]
        df['Gap_Up'] = (df['Low'] > df['High'].shift(1)) & (df['High'].shift(1).notna())
        df['Gap_Down'] = (df['High'] < df['Low'].shift(1)) & (df['Low'].shift(1).notna())
        
        # --- 涨停 / 跌停旗标（板块统计用；无涨跌停制度的标的恒为 0）---
        _r = limit_rate(code, name)
        if _r > 0:
            _pct = df['Close'] / prev_close - 1
            _up_band = (_pct >= _r - LIMIT_TOL_DOWN) & (_pct <= _r + LIMIT_TOL_UP)
            _dn_band = (_pct <= -_r + LIMIT_TOL_DOWN) & (_pct >= -_r - LIMIT_TOL_UP)
            if LIMIT_SEAL_ONLY:
                _sealed_up = (df['High'] - df['Close']).abs() <= df['Close'].abs() * LIMIT_SEAL_TOL
                _sealed_dn = (df['Close'] - df['Low']).abs() <= df['Close'].abs() * LIMIT_SEAL_TOL
                _up_band = _up_band & _sealed_up
                _dn_band = _dn_band & _sealed_dn
            df['Limit_Up'] = _up_band.fillna(False).astype(int)
            df['Limit_Down'] = _dn_band.fillna(False).astype(int)
        else:
            df['Limit_Up'] = 0
            df['Limit_Down'] = 0
        df['Limit_Rate'] = _r

        # --- 计算 MACD ---[cite: 3]
        exp1 = df['Close'].ewm(span=12, adjust=False).mean()
        exp2 = df['Close'].ewm(span=26, adjust=False).mean()
        df['macd_dif'] = exp1 - exp2
        df['macd_dea'] = df['macd_dif'].ewm(span=9, adjust=False).mean()
        df['macd_hist'] = 2 * (df['macd_dif'] - df['macd_dea'])

        # --- 计算 RSI (14) ---[cite: 3]
        delta = df['Close'].diff()
        gain = (delta.where(delta > 0, 0)).rolling(window=14).mean()
        loss = (-delta.where(delta < 0, 0)).rolling(window=14).mean()
        rs = gain / (loss + 1e-6)
        df['rsi'] = 100 - (100 / (1 + rs))
        
        # --- 计算 KDJ ---[cite: 3]
        low_9 = df['Low'].rolling(window=9).min()
        high_9 = df['High'].rolling(window=9).max()
        rsv = (df['Close'] - low_9) / (high_9 - low_9 + 1e-6) * 100
        df['kdj_k'] = rsv.ewm(com=2, adjust=False).mean()
        df['kdj_d'] = df['kdj_k'].ewm(com=2, adjust=False).mean()
        df['kdj_j'] = 3 * df['kdj_k'] - 2 * df['kdj_d']

        # --- 计算 CCI (14) ---[cite: 3]
        tp = (df['High'] + df['Low'] + df['Close']) / 3
        tp_ma = tp.rolling(window=14).mean()
        md = (tp - tp_ma).abs().rolling(window=14).mean()
        df['cci'] = (tp - tp_ma) / (0.015 * md + 1e-6)

        # --- 计算 BIAS (乖离度 6日) ---[cite: 3]
        bi_ma6 = df['Close'].rolling(window=6).mean()
        df['bias'] = ((df['Close'] - bi_ma6) / bi_ma6 * 100).round(2).astype(str) + '%'

        # --- 简易行情类型与多空共振判定示例 ---[cite: 3]
        df['m_type'] = np.where(df['Close'] > df['MA20'], '单边上涨', '单边下跌')
        # 统一口径：SAR改用真实抛物线SAR算法（与5号/6号脚本一致），
        # 不再用"收盘价是否站上5日均线"这种简化代理判断
        df['sar'], df['sar_trend'] = calculate_sar(df['Close'])
        df['sar_stat'] = np.where(df['sar_trend'] > 0, '看涨', '看跌')
        
        df['low_count'] = 0
        df['high_count'] = 0
        df['comp_score'] = 0.0
        df['macd_stat_str'] = np.where(df['macd_hist'] > 0, '多头金叉', '空头死叉')
        
        return df
    except Exception as e:
        print(f"提取异常: {e}")
        return pd.DataFrame()

def get_kline_components(df):
    o, h, l, c = df['Open'], df['High'], df['Low'], df['Close']
    bod = (c - o).abs()
    total_range = h - l
    us = h - np.maximum(o, c)
    ls = np.minimum(o, c) - l
    avg_body = bod.rolling(20, min_periods=5).mean()
    avg_range = total_range.rolling(20, min_periods=5).mean()
    return o, h, l, c, bod, total_range, us, ls, avg_body, avg_range

# ----------------------------- 2. 严格的15种看跌形态 -----------------------------
def detect_hanging_man(df):
    o, h, l, c, bod, total, us, ls, avg_body, avg_range = get_kline_components(df)
    return (bod / (total + 1e-6) < 0.2) & (ls >= 2.5 * bod) & (ls > 0.5 * total) & (us <= 0.1 * total) & (total > 0.8 * avg_range)

def detect_spinning_top_high(df):
    o, h, l, c, bod, total, us, ls, avg_body, avg_range = get_kline_components(df)
    return (bod / (total + 1e-6) < 0.2) & (us >= 1.5 * bod) & (ls >= 1.5 * bod) & (total > avg_range)

def detect_t_line_high(df):
    o, h, l, c, bod, total, us, ls, avg_body, avg_range = get_kline_components(df)
    return (bod / (total + 1e-6) < 0.05) & (ls >= 0.75 * total) & (us <= 0.05 * total) & (total > 0.8 * avg_range)

def detect_bearish_turning_line(df):
    o, h, l, c, bod, total, us, ls, avg_body, avg_range = get_kline_components(df)
    return (bod / (total + 1e-6) < 0.05) & (us >= 0.75 * total) & (ls <= 0.05 * total) & (total > 0.8 * avg_range)

def detect_long_cross_high(df):
    o, h, l, c, bod, total, us, ls, avg_body, avg_range = get_kline_components(df)
    return (bod / (total + 1e-6) < 0.05) & (us >= 0.4 * total) & (ls >= 0.4 * total) & (total > 1.2 * avg_range)

def detect_flat_top(df):
    o, h, l, c, bod, total, us, ls, avg_body, avg_range = get_kline_components(df)
    return ((h - h.shift(1)).abs() / h.shift(1) < 0.001) & (total > 0.8 * avg_range) & (total.shift(1) > 0.8 * avg_range.shift(1)) & (c < o)

def detect_bearish_friend(df):
    o, h, l, c, bod, total, us, ls, avg_body, avg_range = get_kline_components(df)
    return (c.shift(1) > o.shift(1)) & (o > c.shift(1)) & (c < o) & ((c - c.shift(1)).abs() / c.shift(1) < 0.0015) & (bod > 0.5 * avg_body)

def detect_dark_cloud(df):
    o, h, l, c, bod, total, us, ls, avg_body, avg_range = get_kline_components(df)
    mid = (o.shift(1) + c.shift(1)) / 2
    return (c.shift(1) > o.shift(1)) & (o > c.shift(1)) & (c < mid) & (c > o.shift(1)) & (bod.shift(1) > avg_body.shift(1)) & (bod > avg_body)

def detect_downpour(df):
    o, h, l, c, bod, total, us, ls, avg_body, avg_range = get_kline_components(df)
    return (c.shift(1) > o.shift(1)) & (bod.shift(1) > avg_body.shift(1)) & (o < c.shift(1)) & (c < o) & (c < o.shift(1)) & (bod > avg_body)

def detect_bearish_engulfing(df):
    o, h, l, c, bod, total, us, ls, avg_body, avg_range = get_kline_components(df)
    return (c.shift(1) > o.shift(1)) & (c < o) & (o > c.shift(1)) & (c < o.shift(1)) & (bod > 1.2 * bod.shift(1))

def detect_two_crows(df):
    o, h, l, c, bod, total, us, ls, avg_body, avg_range = get_kline_components(df)
    return (c.shift(2) > o.shift(2)) & (o.shift(1) > c.shift(2)) & (c.shift(1) < o.shift(1)) & \
           (o > c.shift(1)) & (c < o) & (c < c.shift(2)) & (o < o.shift(1))

def detect_three_black_crows(df):
    o, h, l, c, bod, total, us, ls, avg_body, avg_range = get_kline_components(df)
    return (c.shift(2) < o.shift(2)) & (c.shift(1) < o.shift(1)) & (c < o) & \
           (o.shift(1) < o.shift(2)) & (o.shift(1) > c.shift(2)) & \
           (o < o.shift(1)) & (o > c.shift(1)) & \
           (c.shift(1) < c.shift(2)) & (c < c.shift(1)) & (bod > 0.8 * avg_body)

def detect_black_three_soldiers_high(df):
    o, h, l, c, bod, total, us, ls, avg_body, avg_range = get_kline_components(df)
    return (c.shift(2) < o.shift(2)) & (c.shift(1) < o.shift(1)) & (c < o) & \
           (bod.shift(2) > avg_body.shift(2)) & (bod.shift(1) < bod.shift(2)) & (bod < bod.shift(1))

def detect_three_falling_continuation(df):
    o, h, l, c, bod, total, us, ls, avg_body, avg_range = get_kline_components(df)
    return (c.shift(2) < o.shift(2)) & (c.shift(1) < o.shift(1)) & (c < o) & \
           (bod.shift(2) > 1.5 * avg_body.shift(2)) & (bod.shift(1) > 1.5 * avg_body.shift(1)) & (bod > 1.5 * avg_body)

def detect_three_falling_stars(df):
    o, h, l, c, bod, total, us, ls, avg_body, avg_range = get_kline_components(df)
    return (c.shift(3) < o.shift(3)) & (bod.shift(3) > 1.5 * avg_body.shift(3)) & \
           (bod.shift(2) < 0.3 * avg_body.shift(2)) & (bod.shift(1) < 0.3 * avg_body.shift(1)) & (bod < 0.3 * avg_body)

# ----------------------------- 3. 严格的看涨形态 -----------------------------
def detect_hammer(df):
    o, h, l, c, bod, total, us, ls, avg_body, avg_range = get_kline_components(df)
    return (bod / (total + 1e-6) < 0.2) & (ls >= 2.5 * bod) & (ls > 0.5 * total) & (us <= 0.1 * total) & (total > 0.8 * avg_range)

def detect_inverted_hammer(df):
    o, h, l, c, bod, total, us, ls, avg_body, avg_range = get_kline_components(df)
    return (bod / (total + 1e-6) < 0.2) & (us >= 2.5 * bod) & (us > 0.5 * total) & (ls <= 0.1 * total) & (total > 0.8 * avg_range)

def detect_bullish_engulfing(df):
    o, h, l, c, bod, total, us, ls, avg_body, avg_range = get_kline_components(df)
    return (c.shift(1) < o.shift(1)) & (c > o) & (o < c.shift(1)) & (c > o.shift(1)) & (bod > 1.2 * bod.shift(1))

def detect_piercing_line(df):
    o, h, l, c, bod, total, us, ls, avg_body, avg_range = get_kline_components(df)
    mid = (o.shift(1) + c.shift(1)) / 2
    return (c.shift(1) < o.shift(1)) & (o < c.shift(1)) & (c > o) & (c > mid) & (c < o.shift(1)) & (bod > avg_body)

def detect_tweezer_bottom(df):
    o, h, l, c, bod, total, us, ls, avg_body, avg_range = get_kline_components(df)
    return ((l - l.shift(1)).abs() / l.shift(1) < 0.001) & (total > 0.8 * avg_range) & (total.shift(1) > 0.8 * avg_range.shift(1)) & (c > o)

def detect_three_white_soldiers(df):
    o, h, l, c, bod, total, us, ls, avg_body, avg_range = get_kline_components(df)
    return (c.shift(2) > o.shift(2)) & (c.shift(1) > o.shift(1)) & (c > o) & \
           (c.shift(1) > c.shift(2)) & (c > c.shift(1)) & \
           (o.shift(1) > o.shift(2)) & (o.shift(1) < c.shift(2)) & \
           (bod.shift(2) > 0.8 * avg_body.shift(2)) & (us < 0.2 * total)

def detect_morning_star(df):
    o, h, l, c, bod, total, us, ls, avg_body, avg_range = get_kline_components(df)
    return (c.shift(2) < o.shift(2)) & (bod.shift(2) > 1.2 * avg_body.shift(2)) & \
           (bod.shift(1) < 0.3 * avg_body.shift(1)) & \
           (c > o) & (bod > 1.2 * avg_body) & (c > (o.shift(2) + c.shift(2))/2)

# ----------------------------- 4. 汇总形态并生成标签 -----------------------------
def extract_pattern_names(df):
    bear = pd.DataFrame(index=df.index)
    bear['吊颈线'] = detect_hanging_man(df)
    bear['螺旋桨'] = detect_spinning_top_high(df)
    bear['T字线'] = detect_t_line_high(df)
    bear['下跌转折线'] = detect_bearish_turning_line(df)
    bear['十字长'] = detect_long_cross_high(df)
    bear['平顶'] = detect_flat_top(df)
    bear['淡友反攻'] = detect_bearish_friend(df)
    bear['乌云盖顶'] = detect_dark_cloud(df)
    bear['倾盆大雨'] = detect_downpour(df)
    bear['顶部穿头破脚'] = detect_bearish_engulfing(df)
    bear['双飞乌鸦'] = detect_two_crows(df)
    bear['三只乌鸦'] = detect_three_black_crows(df)
    bear['顶部黑三兵'] = detect_black_three_soldiers_high(df)
    bear['下跌三连阴'] = detect_three_falling_continuation(df)
    bear['下跌三颗星'] = detect_three_falling_stars(df)

    bull = pd.DataFrame(index=df.index)
    bull['锤子线'] = detect_hammer(df)
    bull['倒锤子线'] = detect_inverted_hammer(df)
    bull['看涨吞没'] = detect_bullish_engulfing(df)
    bull['刺透形态'] = detect_piercing_line(df)
    bull['平底'] = detect_tweezer_bottom(df)
    bull['红三兵'] = detect_three_white_soldiers(df)
    bull['早晨之星'] = detect_morning_star(df)

    # 严格趋势过滤[cite: 3]
    high_price = df['Close'] > df['MA20']
    low_price = df['Close'] < df['MA20']
    
    for col in bear.columns: bear[col] = bear[col] & high_price
    for col in bull.columns: bull[col] = bull[col] & low_price
        
    df['Bearish_Names'] = bear.apply(lambda row: '<br>'.join(row.index[row]), axis=1)
    df['Bullish_Names'] = bull.apply(lambda row: '<br>'.join(row.index[row]), axis=1)
    
    # 构建包含缺口的文字提示[cite: 3]
    df['Gap_Text'] = np.where(df['Gap_Up'], '发现 向上缺口', np.where(df['Gap_Down'], '发现 向下缺口', ''))
    
    return df

# ----------------------------- 5. 绘制并导出单个图表 -----------------------------
def plot_interactive(df, name, code, sector, output_dir, sector_display=None):
    """sector      —— 主板块，用于文件名前缀（一只票只出一张图，不按板块重复出图）
       sector_display —— 标题里展示的完整所属板块，如「CPO / 存储芯片 / 半导体」"""
    # 【性能】只画最近 N 根K线（指标仍用全量历史计算，结果不变）
    if PLOT_RECENT_BARS and len(df) > PLOT_RECENT_BARS:
        df = df.tail(PLOT_RECENT_BARS).copy()

    # 【性能】原写法 `d not in dt_obs` 在 list 上做成员判断是 O(n²)，
    # 20年历史(约5000根K线)光这一步就要几百毫秒；改用 set 后降为 O(n)
    dt_obs = set(df.index.strftime("%Y-%m-%d"))
    dt_all = pd.date_range(start=df.index.min(), end=df.index.max())
    dt_breaks = [d for d in dt_all.strftime("%Y-%m-%d") if d not in dt_obs]

    label = sector_display if sector_display else sector
    label_txt = f"【{label}】 " if label else ""
    title_text = f'{label_txt}{name} ({code}) 高级K线形态与缺口系统'

    fig = make_subplots(rows=4, cols=1, shared_xaxes=True,
                        vertical_spacing=0.03, 
                        row_heights=[0.50, 0.17, 0.17, 0.16],
                        subplot_titles=("", "成交量", "MACD", "RSI"))

    df['Pct_Change'] = (df['Close'].pct_change() * 100).round(2)
    df['Change_Amt'] = (df['Close'] - df['Close'].shift(1)).round(2)
    
    df['Gap_Val'] = np.where(
        df['Gap_Up'], (df['Low'] - df['High'].shift(1)).round(2),
        np.where(df['Gap_Down'], (df['High'] - df['Low'].shift(1)).round(2), 0.0)
    )
    
    df['Dist_Upper'] = (((df['boll_upper'] - df['Close']) / df['Close']) * 100).round(2)
    df['Dist_Mid'] = (((df['MA20'] - df['Close']) / df['Close']) * 100).round(2)
    df['Dist_Lower'] = (((df['Close'] - df['boll_lower']) / df['Close']) * 100).round(2)
    
    
    fig.add_trace(go.Candlestick(
        x=df.index, open=df['Open'], high=df['High'], low=df['Low'], close=df['Close'],
        increasing_line_color='#FF3333', increasing_fillcolor='#FF3333',
        decreasing_line_color='#00CC00', decreasing_fillcolor='#00CC00',
        name='K线',
        customdata=df[[
            'VP_Analysis', 'Volume', 'Gap_Text', 'Pct_Change', 
            'Change_Amt', 'Gap_Val', 'Dist_Upper', 'Dist_Mid', 'Dist_Lower'
        ]],
        hovertemplate=(
            '<b>日期:</b> %{x}<br>'
            '<b>开盘:</b> %{open:.2f}<br>'
            '<b>收盘:</b> %{close:.2f} (<b>变化: %{customdata[4]:+.2f} | %{customdata[3]:+.2f}%</b>)<br>'
            '<b>成交量:</b> %{customdata[1]:,.0f}<br>'
            '<b>量价分析:</b> %{customdata[0]}<br>'
            '<b>缺口值:</b> %{customdata[5]:.2f}<br>'
            '<b>距上轨:</b> %{customdata[6]:+.2f}% | '
            '<b>距中轨:</b> %{customdata[7]:+.2f}% | '
            '<b>距下轨:</b> %{customdata[8]:+.2f}%<br>'
            '<span style="color:#E5C07B"><b>%{customdata[2]}</b></span><extra></extra>'
        )
    ), row=1, col=1)

    # 【性能】fig.add_shape 每调用一次都会重新校验整个 layout.shapes 元组，
    # 循环添加是 O(k²)：500个缺口要 16 秒。改为先攒 list，最后一次性写入（0.07秒）
    # row=1,col=1 对应 xref='x' / yref='y'，与逐个 add_shape 的效果完全一致
    gap_shapes = []
    highs, lows = df['High'].values, df['Low'].values
    idx_arr = df.index

    for loc in np.flatnonzero(df['Gap_Up'].values):
        if loc > 0:
            gap_shapes.append(dict(type="rect", xref="x", yref="y",
                                   x0=idx_arr[loc-1], x1=idx_arr[loc],
                                   y0=highs[loc-1], y1=lows[loc],
                                   fillcolor="rgba(0, 149, 255, 0.5)",
                                   line=dict(color="#1E90FF", width=2),
                                   layer="below"))

    for loc in np.flatnonzero(df['Gap_Down'].values):
        if loc > 0:
            gap_shapes.append(dict(type="rect", xref="x", yref="y",
                                   x0=idx_arr[loc-1], x1=idx_arr[loc],
                                   y0=lows[loc-1], y1=highs[loc],
                                   fillcolor="rgba(0, 110, 220, 0.4)",
                                   line=dict(color="#00BFFF", width=2),
                                   layer="below"))

    if gap_shapes:
        fig.update_layout(shapes=gap_shapes)

    fig.add_trace(go.Scatter(x=df.index, y=df['boll_upper'], name="BOLL上轨", line=dict(dash='dash', color='rgba(128,128,128,0.5)'), hoverinfo='skip'), row=1, col=1)
    fig.add_trace(go.Scatter(x=df.index, y=df['boll_lower'], name="BOLL下轨", line=dict(dash='dash', color='rgba(128,128,128,0.5)'), hoverinfo='skip'), row=1, col=1)
    fig.add_trace(go.Scatter(x=df.index, y=df['MA5'], mode='lines', line=dict(color='#E5C07B', width=1), name='MA5'), row=1, col=1)
    fig.add_trace(go.Scatter(x=df.index, y=df['MA10'], mode='lines', line=dict(color='#61AFEF', width=1), name='MA10'), row=1, col=1)
    fig.add_trace(go.Scatter(x=df.index, y=df['MA20'], mode='lines', line=dict(color='#C678DD', width=1.5), name='MA20(中轨)'), row=1, col=1)

    bearish_mask = df['Bearish_Names'] != ''
    if bearish_mask.any():
        fig.add_trace(go.Scatter(
            x=df.index[bearish_mask], y=df['High'][bearish_mask] * 1.015,
            mode='markers', marker=dict(symbol='triangle-down', size=12, color='yellow', line=dict(color='black', width=1)),
            name='看跌信号', customdata=df['Bearish_Names'][bearish_mask],
            hovertemplate='<b>日期:</b> %{x}<br><b>看跌预警:</b><br><span style="color:yellow">%{customdata}</span><extra></extra>'
        ), row=1, col=1)

    bullish_mask = df['Bullish_Names'] != ''
    if bullish_mask.any():
        fig.add_trace(go.Scatter(
            x=df.index[bullish_mask], y=df['Low'][bullish_mask] * 0.985,
            mode='markers', marker=dict(symbol='triangle-up', size=12, color='#FF00FF', line=dict(color='black', width=1)),
            name='看涨信号', customdata=df['Bullish_Names'][bullish_mask],
            hovertemplate='<b>日期:</b> %{x}<br><b>看涨预警:</b><br><span style="color:#FF00FF">%{customdata}</span><extra></extra>'
        ), row=1, col=1)

    vol_colors = ['#FF3333' if df.loc[i, 'Close'] >= df.loc[i, 'Open'] else '#00CC00' for i in df.index]
    fig.add_trace(go.Bar(
        x=df.index, y=df['Volume'], name='成交量', marker_color=vol_colors,
        customdata=df['VP_Analysis'], hovertemplate='<b>量价:</b> %{customdata}<extra></extra>', showlegend=False
    ), row=2, col=1)
    
    max_date = df.index.max()
    min_date = df.index.min()
    pad_date = max_date + pd.Timedelta(days=2)
    
    last_row = df.iloc[-1]
    last_vol = last_row['Volume']

    fig.add_shape(
        type="line", x0=df.index.min(), x1=pad_date, y0=last_vol, y1=last_vol,
        xref="x", yref="y2", line=dict(color="#61AFEF", width=1.5, dash="dot"), row=2, col=1
    )
    
    fig.add_trace(go.Scatter(x=df.index, y=df['macd_dif'], name='MACD DIF', line=dict(color='#E5C07B', width=1)), row=3, col=1)
    fig.add_trace(go.Scatter(x=df.index, y=df['macd_dea'], name='MACD DEA', line=dict(color='#61AFEF', width=1)), row=3, col=1)
    
    macd_colors = ['#FF3333' if val >= 0 else '#00CC00' for val in df['macd_hist']]
    fig.add_trace(go.Bar(x=df.index, y=df['macd_hist'], name='MACD Hist', marker_color=macd_colors, showlegend=False), row=3, col=1)

    fig.add_trace(go.Scatter(x=df.index, y=df['rsi'], name='RSI(14)', line=dict(color='#C678DD', width=1)), row=4, col=1)
    fig.add_hline(y=70, line_dash="dash", line_color="rgba(255, 51, 51, 0.5)", row=4, col=1)
    fig.add_hline(y=30, line_dash="dash", line_color="rgba(0, 204, 0, 0.5)", row=4, col=1)
    
    fig.update_layout(
        template='plotly_dark', title=dict(text=title_text, x=0.5, font=dict(size=18)),
        xaxis_rangeslider_visible=False, hovermode='x unified', height=1100, 
        margin=dict(l=100, r=100, t=80, b=50), 
        legend=dict(orientation='h', yanchor='bottom', y=1.02, xanchor='right', x=1),
        plot_bgcolor='#161616', paper_bgcolor='#161616'
    )

    fig.update_xaxes(rangebreaks=[dict(values=dt_breaks)])
    # 横轴统一显示 2026-03-06 这种格式，不用 Plotly 默认的英文月份缩写
    fig.update_xaxes(tickformat="%Y-%m-%d", hoverformat="%Y-%m-%d")
    fig.update_yaxes(title_text='价格点位', gridcolor='#2a2a2a', row=1, col=1)
    fig.update_yaxes(title_text='成交量', gridcolor='#2a2a2a', row=2, col=1)
    fig.update_yaxes(title_text='MACD', gridcolor='#2a2a2a', row=3, col=1)
    fig.update_yaxes(title_text='RSI', gridcolor='#2a2a2a', row=4, col=1)
    fig.update_xaxes(gridcolor='#2a2a2a', row=4, col=1)

    def get_y_range(months=0, years=0):
        if months: start = max_date - pd.DateOffset(months=months)
        elif years: start = max_date - pd.DateOffset(years=years)
        else: start = min_date
            
        mask = df.index >= start
        if not mask.any():
            return [df['Low'].min() * 0.95, df['High'].max() * 1.05]
            
        local_min = df.loc[mask, 'Low'].min()
        local_max = df.loc[mask, 'High'].max()
        padding = (local_max - local_min) * 0.05
        if padding == 0: padding = local_max * 0.05
        return [local_min - padding, local_max + padding]

    buttons = [
        dict(label='1月', method='relayout', args=[{'xaxis.range': [max_date - pd.DateOffset(months=1), pad_date], 'yaxis.range': get_y_range(months=1)}]),
        dict(label='3月', method='relayout', args=[{'xaxis.range': [max_date - pd.DateOffset(months=3), pad_date], 'yaxis.range': get_y_range(months=3)}]),
        dict(label='6月', method='relayout', args=[{'xaxis.range': [max_date - pd.DateOffset(months=6), pad_date], 'yaxis.range': get_y_range(months=6)}]),
        dict(label='1年', method='relayout', args=[{'xaxis.range': [max_date - pd.DateOffset(years=1), pad_date], 'yaxis.range': get_y_range(years=1)}]),
        dict(label='全部', method='relayout', args=[{'xaxis.range': [min_date, pad_date], 'yaxis.range': get_y_range()}])
    ]

    all_shapes = list(fig.layout.shapes)
    gap_buttons = [
        dict(label='显示缺口', method='relayout', args=[{'shapes': all_shapes}]),
        dict(label='隐藏缺口', method='relayout', args=[{'shapes': []}])
    ]

    fig.update_layout(
        updatemenus=[
            dict(buttons=buttons, direction='down', showactive=True, x=0.01, y=1.1, xanchor='left', yanchor='top', bgcolor='#2a2a2a', font=dict(color='white'), bordercolor='#444'),
            dict(buttons=gap_buttons, direction='down', showactive=True, x=0.15, y=1.1, xanchor='left', yanchor='top', bgcolor='#2a2a2a', font=dict(color='white'), bordercolor='#444')
        ]
    )
    
    fig.update_xaxes(range=[max_date - pd.DateOffset(years=1), pad_date])
    fig.update_yaxes(range=get_y_range(years=1), row=1, col=1)
    
    curr_close = last_row['Close']
    dist_to_upper = ((last_row['boll_upper'] - curr_close) / curr_close) * 100
    dist_to_lower = ((curr_close - last_row['boll_lower']) / curr_close) * 100
    
    rsi_val = last_row.get('rsi', 50)
    kdj_j = last_row.get('kdj_j', 50)
    cci_val = last_row.get('cci', 0)
    bias_val = last_row.get('bias', '0.00%')
    sar_stat = last_row.get('sar_stat', '中性')
    m_type = last_row.get('m_type', '震荡行情')
    macd_stat_str = last_row.get('macd_stat_str', '金叉')
    macd_hist_val = last_row.get('macd_hist', 0)

    if rsi_val <= 10: rsi_color = "#88f933"
    elif rsi_val <= 30: rsi_color = "#357c00"
    elif rsi_val >= 90: rsi_color = "#f93333"
    elif rsi_val >= 70: rsi_color = "#7c0000"
    else: rsi_color = "#ffffff"

    if kdj_j <= -10: kdj_color = "#88f933"
    elif kdj_j <= 10: kdj_color = "#357c00"
    elif kdj_j >= 110: kdj_color = "#f93333"
    elif kdj_j >= 90: kdj_color = "#7c0000"
    else: kdj_color = "#ffffff"

    if cci_val <= -200: cci_color = "#88f933"
    elif cci_val <= -100: cci_color = "#357c00"
    elif cci_val >= 200: cci_color = "#f93333"
    elif cci_val >= 100: cci_color = "#7c0000"
    else: cci_color = "#ffffff"

    try: bias_num = float(str(bias_val).replace('%',''))
    except: bias_num = 0

    if bias_num <= -8: bias_color = "#88f933"
    elif bias_num <= -4: bias_color = "#357c00"
    elif bias_num >= 8: bias_color = "#f93333"
    elif bias_num >= 4: bias_color = "#7c0000"
    else: bias_color = "#ffffff"

    sar_color = "#c0392b" if sar_stat == "看涨" else ("#27ae60" if sar_stat == "看跌" else "#ffffff")

    if m_type == "单边上涨": type_color = "#c0392b" if rsi_val > 70 else "#f5b7b1"
    elif m_type in ["底部反转", "底部修复"]: type_color = "#a9dfbf"  
    elif m_type == "单边下跌": type_color = "#27ae60"  
    elif m_type == "窄幅震荡": type_color = "#FFD700"  
    else: type_color = "#ffffff"
        
    close_color = "#17BECF"
    
    if "金叉" in macd_stat_str or "多头" in macd_stat_str:
        macd_txt_color = "#ff5555" if macd_hist_val > 0 else "#ffcccc"
    else:
        macd_txt_color = "#55ff55" if macd_hist_val < 0 else "#ccffcc"
    
    dashboard_text = (
        f"<span style='color:#FFFFFF; font-weight:bold;'>单位净值: </span>"
        f"<span style='color:{close_color}; font-weight:bold;'>{curr_close:.4f}</span><br>"
        f"<span style='color:#FFFFFF; font-weight:bold;'>轨道空间: </span>"
        f"<span style='color:#ff5555; font-weight:bold;'>距上轨 {dist_to_upper:+.2f}%</span> | "
        f"<span style='color:#55ff55; font-weight:bold;'>距下轨 {dist_to_lower:+.2f}%</span><br>"
        f"<span style='color:#FFFFFF; font-weight:bold;'>当前行情: </span>"
        f"<span style='color:{type_color}; font-weight:bold;'>{m_type}</span><br>"
        f"<span style='color:#FFFFFF; font-weight:bold;'>MACD形态: </span>"
        f"<span style='color:{macd_txt_color}; font-weight:bold;'>{macd_stat_str} ({macd_hist_val:+.4f})</span><br>"
        f"<span style='color:#FFFFFF; font-weight:bold;'>RSI 指标: </span>"
        f"<span style='color:{rsi_color}; font-weight:bold;'>{rsi_val:.1f}</span><br>"
        f"<span style='color:#FFFFFF; font-weight:bold;'>乖 离 度 : </span>"
        f"<span style='color:{bias_color}; font-weight:bold;'>{bias_val}</span><br>"
        f"<span style='color:#FFFFFF; font-weight:bold;'>KDJ(J)值 : </span>"
        f"<span style='color:{kdj_color}; font-weight:bold;'>{kdj_j:.2f}</span><br>"
        f"<span style='color:#FFFFFF; font-weight:bold;'>CCI 指标 : </span>"
        f"<span style='color:{cci_color}; font-weight:bold;'>{cci_val:.2f}</span><br>"
        f"<span style='color:#FFFFFF; font-weight:bold;'>SAR 状态 : </span>"
        f"<span style='color:{sar_color}; font-weight:bold;'>{sar_stat}</span>"
    )

    fig.add_annotation(
        text=dashboard_text, xref="paper", yref="paper", x=0.01, y=0.98,
        showarrow=False, align="left", bgcolor="rgba(22, 22, 22, 0.85)",
        bordercolor="#444", borderwidth=1, borderpad=6, font=dict(size=11, color="#ffffff")
    )
    
    safe = lambda x: "".join(ch for ch in str(x) if ch not in '\\/:*?"<>|').strip()
    file_prefix = f"[{safe(sector)}]_" if sector else ""
    name, code = safe(name), safe(code)
    filepath = os.path.join(output_dir, f"{file_prefix}{name}_{code}_形态与缺口分析.html")
    # 【性能】默认 include_plotlyjs=True 会把 3MB 的 plotly.js 内嵌进每个 HTML，
    # 90个文件就是 400MB+ 的磁盘写入（Windows 上还要过一遍杀软扫描）。
    # 'directory' 只在输出目录放一份 plotly.min.js，各图表引用它，依然可离线打开
    fig.write_html(filepath, include_plotlyjs=PLOTLY_JS_MODE,
                   full_html=True, auto_open=False, config={'responsive': True})
    # Plotly 默认不写 viewport meta，手机打开会按桌面宽度渲染再整体缩小；补一行
    with open(filepath, 'r', encoding='utf-8') as f:
        _html = f.read()
    if 'name="viewport"' not in _html:
        _html = _html.replace(
            '<head>',
            '<head>\n    <meta name="viewport" content="width=device-width, initial-scale=1">',
            1
        )
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(_html)
    print(f"✅ 成功生成图表: {filepath}")

# ----------------------------- 5.5 板块股票池：读取 / 勾选 / 去重 -----------------------------
def read_csv_smart(path, **kw):
    """CSV 编码兼容：utf-8-sig → gbk → gb18030"""
    for enc in ("utf-8-sig", "gbk", "gb18030"):
        try:
            return pd.read_csv(path, encoding=enc, **kw)
        except UnicodeDecodeError:
            continue
    return pd.read_csv(path, encoding="utf-8", errors="ignore", **kw)


def load_board_pool(path=BOARD_CSV):
    """读取 ztjj_board_stocks.csv，标准化为 板块/类别/代码/名称/持仓比例 五列"""
    if not os.path.exists(path):
        print(f"⚠️ 未找到板块股票池文件：{path}")
        return pd.DataFrame(columns=['板块', '类别', 'code', 'name', 'ratio'])

    df = read_csv_smart(path, dtype={'股票代码': str})
    need = ['板块名称', '板块类别', '股票代码', '股票名称']
    miss = [c for c in need if c not in df.columns]
    if miss:
        print(f"⚠️ {path} 缺少列 {miss}，无法解析")
        return pd.DataFrame(columns=['板块', '类别', 'code', 'name', 'ratio'])

    ratio_col = '基金持有比例(%)'
    out = pd.DataFrame({
        '板块': df['板块名称'].astype(str).str.strip(),
        '类别': df['板块类别'].astype(str).str.strip(),
        # 代码在 CSV 里是数字，000001 会被存成 1，统一补足 6 位
        'code': df['股票代码'].astype(str).str.strip().str.zfill(6),
        'name': df['股票名称'].astype(str).str.strip(),
        'ratio': pd.to_numeric(df[ratio_col], errors='coerce') if ratio_col in df.columns else np.nan,
    })
    out = out[out['code'].str.len() == 6]
    return out.dropna(subset=['板块', 'code']).drop_duplicates(subset=['板块', 'code'])


def board_catalog(pool):
    """{类别: [(板块名, 成分股数), ...]}，按成分股数倒序"""
    cat = {}
    for c, g in pool.groupby('类别'):
        items = g.groupby('板块')['code'].nunique().sort_values(ascending=False)
        cat[c] = [(k, int(v)) for k, v in items.items()]
    return cat


def preview_count(pool, boards, top_n):
    """预估去重后的股票数（对话框里实时提示用）"""
    if not boards:
        return 0
    sub = pool[pool['板块'].isin(boards)]
    if top_n and top_n > 0:
        sub = (sub.sort_values('ratio', ascending=False, na_position='last')
                  .groupby('板块').head(top_n))
    return int(sub['code'].nunique())


def _defaults_for(catalog, category, defaults=DEFAULT_BOARDS):
    """默认板块里，属于当前类别且确实存在的那些"""
    if category == '全部':
        names = {n for lst in catalog.values() for n, _ in lst}
    else:
        names = {n for n, _ in catalog.get(category, [])}
    # 大小写容错：配置里写 cpo 也能命中 CPO
    lower = {n.lower(): n for n in names}
    hit = []
    for d in defaults:
        real = lower.get(str(d).strip().lower())
        if real and real not in hit:
            hit.append(real)
    return hit


def select_boards_ui(pool, catalog):
    """弹出勾选对话框；无图形界面则返回 None 交给命令行版本"""
    try:
        import tkinter as tk
        from tkinter import ttk, messagebox
    except Exception:
        return None

    try:
        root = tk.Tk()
    except Exception:
        return None   # 服务器/无显示环境

    result = {}
    root.title("选择要分析的板块")
    root.geometry("1180x720")
    root.minsize(820, 560)           # 再窄底部控件就要挤了

    state = dict(category=DEFAULT_BOARD_CATEGORY if DEFAULT_BOARD_CATEGORY in ('概念', '行业', '全部') else '概念')
    vars_map = {}          # 板块名 -> BooleanVar（跨类别常驻，切换类别不丢勾选）

    top = ttk.Frame(root, padding=(12, 10, 12, 4))
    top.pack(fill='x')
    ttk.Label(top, text="板块类别：").pack(side='left')

    cat_var = tk.StringVar(value=state['category'])
    kw_var = tk.StringVar()
    topn_var = tk.StringVar(value=str(DEFAULT_TOP_N_PER_BOARD))
    ovs_var = tk.BooleanVar(value=INCLUDE_OVERSEAS)
    legacy_var = tk.BooleanVar(value=INCLUDE_LEGACY_DOMESTIC)

    for c in ('概念', '行业', '全部'):
        ttk.Radiobutton(top, text=c, value=c, variable=cat_var,
                        command=lambda: switch_category()).pack(side='left', padx=6)

    ttk.Label(top, text="   搜索：").pack(side='left')
    ent = ttk.Entry(top, textvariable=kw_var, width=18)
    ent.pack(side='left')
    kw_var.trace_add('write', lambda *a: refresh_list())

    box = ttk.Frame(root, padding=(12, 4))
    box.pack(fill='both', expand=True)
    canvas = tk.Canvas(box, highlightthickness=0)
    bar = ttk.Scrollbar(box, orient='vertical', command=canvas.yview)
    inner = ttk.Frame(canvas)
    inner.bind('<Configure>', lambda e: canvas.configure(scrollregion=canvas.bbox('all')))
    # 记住这个 window id：窗口变宽时要把内层一起撑开，否则右边全是空白
    win_id = canvas.create_window((0, 0), window=inner, anchor='nw')
    canvas.configure(yscrollcommand=bar.set)
    canvas.pack(side='left', fill='both', expand=True)
    bar.pack(side='right', fill='y')
    canvas.bind_all('<MouseWheel>', lambda e: canvas.yview_scroll(int(-e.delta / 120), 'units'))

    COL_W = 210                      # 单个复选框大致占多宽（含板块名 + 成分股数）
    layout = {'cols': 0}             # 当前列数，只有变了才重排，避免拖动窗口时狂刷

    def apply_grid(width):
        """按当前宽度把列宽平均分配（用 minsize 而不是 weight/uniform：
        后者在窗口变窄时会把列压得比文字还短，出现文字被截断）"""
        cols = layout['cols'] or 4
        cw = max(140, (width - 16) // cols)
        for c in range(12):
            inner.grid_columnconfigure(c, weight=0, minsize=cw if c < cols else 0)

    def on_canvas_resize(e):
        canvas.itemconfigure(win_id, width=e.width)          # 内层跟着画布一样宽
        cols = max(1, min(10, e.width // COL_W))             # 列数随宽度自适应
        if cols != layout['cols']:
            layout['cols'] = cols
            refresh_list()                                   # 列数变了才重排
        else:
            apply_grid(e.width)                              # 只是宽度微调，改列宽即可
    canvas.bind('<Configure>', on_canvas_resize)

    # 底部分两行：第一行是统计文字（会自动换行，不再被右侧控件挤掉），第二行才是控件
    bottom = ttk.Frame(root, padding=(12, 6, 12, 10))
    bottom.pack(fill='x')
    tip = ttk.Label(bottom, text="", anchor='w', justify='left')
    tip.pack(fill='x', pady=(0, 6))
    # 控件再拆成两行：一行选项、一行操作按钮。挤在一行时总宽要 1250px，
    # 窗口一窄右边的“开始分析/取消”就会被裁掉
    opts = ttk.Frame(bottom)
    opts.pack(fill='x', pady=(0, 6))
    acts = ttk.Frame(bottom)
    acts.pack(fill='x')

    def on_root_resize(e=None):
        w = root.winfo_width()
        if w > 100 and abs(w - layout.get('w', 0)) > 20:
            layout['w'] = w
            tip.config(wraplength=w - 40)                    # 文字按窗口宽度换行
    root.bind('<Configure>', on_root_resize)

    def current_items():
        if cat_var.get() == '全部':
            items = []
            for c in ('概念', '行业'):
                items += [(n, k, c) for n, k in catalog.get(c, [])]
            for c in catalog:
                if c not in ('概念', '行业'):
                    items += [(n, k, c) for n, k in catalog[c]]
        else:
            items = [(n, k, cat_var.get()) for n, k in catalog.get(cat_var.get(), [])]
        kw = kw_var.get().strip().lower()
        if kw:
            items = [i for i in items if kw in i[0].lower()]
        return items

    def selected():
        return [n for n, v in vars_map.items() if v.get()]

    def update_tip(*a):
        try:
            n = int(topn_var.get() or 0)
        except ValueError:
            n = 0
        sel = selected()
        visible = {nm for nm, _, _ in current_items()}
        other = [x for x in sel if x not in visible]
        extra = f"｜其中 {len(other)} 个不在当前视图：{'、'.join(other[:6])}" if other else ""
        tip.config(text=f"已选 {len(sel)} 个板块 ≈ {preview_count(pool, sel, n)} 只股票（已按代码去重）{extra}")

    def refresh_list():
        for w in inner.winfo_children():
            w.destroy()
        items = current_items()
        cols = layout['cols'] or 4
        for i, (nm, cnt, cat) in enumerate(items):
            if nm not in vars_map:
                vars_map[nm] = tk.BooleanVar(value=False)
            txt = f"{nm}（{cnt}）" if cat_var.get() != '全部' else f"{nm}·{cat}（{cnt}）"
            cb = ttk.Checkbutton(inner, text=txt, variable=vars_map[nm], command=update_tip)
            cb.grid(row=i // cols, column=i % cols, sticky='w', padx=6, pady=3)
        # 各列等宽平分，窗口拉宽时复选框会铺满，而不是全挤在左边
        apply_grid(canvas.winfo_width() or 900)
        update_tip()

    def switch_category():
        # 切类别时把该类别的默认板块自动勾上（只在第一次进入该类别时补勾）
        for nm in _defaults_for(catalog, cat_var.get()):
            vars_map.setdefault(nm, tk.BooleanVar(value=True))
        refresh_list()

    def check_all(flag):
        for nm, _, _ in current_items():
            vars_map.setdefault(nm, tk.BooleanVar(value=False)).set(flag)
        update_tip()

    def restore_default():
        # 默认清单是跨类别的（有色金属/电力/白酒/银行属“行业”），一次性全部勾上，
        # 这样即便当前只在看“概念”，那几个行业板块也已在选中列表里
        for v in vars_map.values():
            v.set(False)
        for nm in _defaults_for(catalog, '全部'):
            vars_map.setdefault(nm, tk.BooleanVar(value=False)).set(True)
        update_tip()

    def confirm():
        sel = selected()
        if not sel:
            messagebox.showwarning("提示", "请至少勾选一个板块")
            return
        try:
            n = max(0, int(topn_var.get() or 0))
        except ValueError:
            n = 0
        result.update(boards=sel, top_n=n, category=cat_var.get(),
                      overseas=ovs_var.get(), legacy=legacy_var.get())
        root.destroy()

    # —— 第一行：选择相关的辅助操作 + 取数范围选项 ——
    ttk.Button(opts, text="恢复默认", command=restore_default).pack(side='left', padx=(0, 4))
    ttk.Button(opts, text="全选", command=lambda: check_all(True)).pack(side='left', padx=4)
    ttk.Button(opts, text="清空", command=lambda: check_all(False)).pack(side='left', padx=4)

    ttk.Checkbutton(opts, text="并入 target_stocks 国内", variable=legacy_var).pack(side='right', padx=(8, 0))
    ttk.Checkbutton(opts, text="包含国外标的", variable=ovs_var).pack(side='right', padx=8)
    ttk.Entry(opts, textvariable=topn_var, width=5).pack(side='right')
    ttk.Label(opts, text="每板块取前 N 只：").pack(side='right', padx=(16, 4))
    topn_var.trace_add('write', update_tip)

    # —— 第二行：说明文字 + 主操作按钮 ——
    ttk.Label(acts, text="N 按基金持仓比例排序，0 = 全部", foreground="#666").pack(side='left')
    ttk.Button(acts, text="取消", command=root.destroy).pack(side='right', padx=(4, 0))
    ttk.Button(acts, text="开始分析", command=confirm).pack(side='right', padx=4)

    restore_default()
    refresh_list()
    root.mainloop()
    return result or None


def select_boards_console(pool, catalog):
    """无图形界面时的命令行选择：直接回车 = 用默认"""
    print("\n请选择板块类别： 1) 概念(默认)  2) 行业  3) 全部")
    ans = input("输入序号后回车（直接回车用默认）: ").strip()
    category = {'2': '行业', '3': '全部'}.get(ans, DEFAULT_BOARD_CATEGORY)

    items = []
    if category == '全部':
        for c in catalog:
            items += [(n, k) for n, k in catalog[c]]
    else:
        items = catalog.get(category, [])

    default_sel = _defaults_for(catalog, '全部')      # 默认清单跨类别，与对话框一致
    print(f"\n【{category}】共 {len(items)} 个板块：")
    for i, (n, k) in enumerate(items, 1):
        mark = '*' if n in default_sel else ' '
        print(f"  {mark}{i:>3}. {n}({k})", end='\n' if i % 4 == 0 else '')
    print(f"\n（带 * 为默认勾选：{'、'.join(default_sel)}）")

    raw = input("输入板块序号或名称，逗号分隔（直接回车用默认）: ").strip()
    if not raw:
        boards = default_sel
    else:
        boards, names = [], [n for n, _ in items]
        for p in raw.replace('，', ',').split(','):
            p = p.strip()
            if not p:
                continue
            if p.isdigit() and 1 <= int(p) <= len(names):
                boards.append(names[int(p) - 1])
            elif p in names:
                boards.append(p)
        boards = list(dict.fromkeys(boards)) or default_sel

    raw_n = input(f"每板块取前 N 只(按基金持仓比例, 0=全部, 默认 {DEFAULT_TOP_N_PER_BOARD}): ").strip()
    try:
        top_n = max(0, int(raw_n)) if raw_n else DEFAULT_TOP_N_PER_BOARD
    except ValueError:
        top_n = DEFAULT_TOP_N_PER_BOARD

    return dict(boards=boards, top_n=top_n, category=category,
                overseas=INCLUDE_OVERSEAS, legacy=INCLUDE_LEGACY_DOMESTIC)


def choose_boards(pool):
    """统一入口：优先图形勾选，失败退回命令行，再失败用配置默认值"""
    catalog = board_catalog(pool)
    if not catalog:
        return dict(boards=[], top_n=0, category=DEFAULT_BOARD_CATEGORY,
                    overseas=INCLUDE_OVERSEAS, legacy=INCLUDE_LEGACY_DOMESTIC)

    if USE_SELECTION_UI:
        try:
            res = select_boards_ui(pool, catalog)
        except Exception as e:
            print(f"⚠️ 勾选窗口异常({e})，改用命令行选择")
            res = None
        if res:
            return res
        print("（未检测到图形界面或窗口被关闭，改用命令行选择）")
        try:
            return select_boards_console(pool, catalog)
        except (EOFError, KeyboardInterrupt):
            pass

    return dict(boards=_defaults_for(catalog, '全部'),
                top_n=DEFAULT_TOP_N_PER_BOARD, category=DEFAULT_BOARD_CATEGORY,
                overseas=INCLUDE_OVERSEAS, legacy=INCLUDE_LEGACY_DOMESTIC)


def build_domestic_targets(pool, boards, top_n):
    """选中的板块 → {股票代码: 任务}，同一只票在多个板块里只保留一条、板块合并进 boards"""
    sub = pool[pool['板块'].isin(boards)].copy()
    if sub.empty:
        return {}
    if top_n and top_n > 0:
        sub = (sub.sort_values('ratio', ascending=False, na_position='last')
                  .groupby('板块').head(top_n))

    # 按用户勾选顺序排板块，保证“主板块”（文件名前缀）稳定可预期
    order = {b: i for i, b in enumerate(boards)}
    sub['_o'] = sub['板块'].map(order).fillna(9999)
    sub = sub.sort_values(['code', '_o'])

    targets = {}
    for code, g in sub.groupby('code', sort=False):
        g = g.sort_values('_o')
        targets[code] = dict(
            code=code, name=g['name'].iloc[0], region='国内', kind='板块',
            boards=[{'板块': r['板块'], '类别': r['类别']} for _, r in g.iterrows()],
        )
    return targets


def build_legacy_targets(path, name2cat, include_overseas=True, include_domestic=False):
    """target_stocks.csv：国外标的保持原样；国内部分默认不并入"""
    if not os.path.exists(path):
        if include_overseas:
            print(f"⚠️ 未找到 {path}，跳过国外标的")
        return {}

    df = read_csv_smart(path, dtype={'code': str})
    if 'global' not in df.columns:
        df['global'] = '国内'
    df['global'] = df['global'].fillna('国内').astype(str).str.strip()
    df.loc[~df['global'].isin(['国内', '国外']), 'global'] = '国内'

    keep = []
    if include_overseas:
        keep.append('国外')
    if include_domestic:
        keep.append('国内')
    df = df[df['global'].isin(keep)]

    targets = {}
    for _, r in df.iterrows():
        code = str(r['code']).strip()
        sector = '' if pd.isna(r.get('sector')) else str(r.get('sector')).strip()
        # 板块名对齐 ztjj（只为“全球合并”能跟 A 股同名板块合并，标的本身不变）
        board = OVERSEAS_BOARD_ALIAS.get(sector, sector) if r['global'] == '国外' else sector
        cat = name2cat.get(board, OVERSEAS_FALLBACK_CATEGORY if board else '未分类')
        t = targets.setdefault(code, dict(code=code, name=str(r['name']).strip(),
                                          region=r['global'], kind='板块', boards=[]))
        if board and board not in [b['板块'] for b in t['boards']]:
            t['boards'].append({'板块': board, '类别': cat})
    return targets


def merge_targets(*groups):
    """跨来源按代码去重：同一只票的板块合并，数据只抓一次、图只画一张"""
    merged = {}
    for g in groups:
        for code, t in g.items():
            if code not in merged:
                merged[code] = dict(t, boards=list(t['boards']))
            else:
                exist = {b['板块'] for b in merged[code]['boards']}
                for b in t['boards']:
                    if b['板块'] not in exist:
                        merged[code]['boards'].append(b)
                        exist.add(b['板块'])
    return merged


# ----------------------------- 6. 并发任务单元 -----------------------------
def normalize_code(code_raw: str) -> str:
    """统一代码格式：宽基指数原样保留，A股补齐6位并加交易所前缀，美股原样"""
    code = str(code_raw)
    if code in BROAD_INDICES:
        return code
    if code.isdigit():
        code = code.zfill(6)
    if not code.startswith(("sh", "sz", "bj")):
        # ztjj 里出现过 689(科创CDR) / 302(深主板新号段) / 920(北交所新号段)，一并覆盖
        if code.startswith(("600", "601", "603", "605", "688", "689", "50", "51")):
            code = "sh" + code
        elif code.startswith(("000", "001", "002", "003", "300", "301", "302", "399")):
            code = "sz" + code
        elif code.startswith(("43", "83", "87", "92")):
            code = "bj" + code
    return code


def build_signal_frame(df, kind, region, boards, code):
    """把单只标的的形态结果压成板块统计所需的信号表

    boards: [{'板块': 名称, '类别': 行业/概念/宽基/...}, ...]
    一只票可能同属多个板块 —— 形态只算一次，这里把同一份结果分发到它的每个板块，
    这样既不重复抓数/绘图，各板块的统计口径又是完整的
    """
    base = pd.DataFrame(index=df.index)
    base['看涨'] = (df['Bullish_Names'] != '').astype(int)
    base['看跌'] = (df['Bearish_Names'] != '').astype(int)
    base['向上缺口'] = df['Gap_Up'].astype(int)
    base['向下缺口'] = df['Gap_Down'].astype(int)
    base['量价看多'] = df['VP_Bull'].astype(int)
    base['量价看空'] = df['VP_Bear'].astype(int)
    # 涨跌停：只做独立统计列，不并入利好/利空
    # （涨停本身就是当日大涨的结果，塞进多空计数会与"向上缺口/放量上涨"严重共线，
    #   把同一件事重复计三遍，多空比会被机械放大）
    base['涨停'] = df['Limit_Up'].astype(int) if 'Limit_Up' in df.columns else 0
    base['跌停'] = df['Limit_Down'].astype(int) if 'Limit_Down' in df.columns else 0
    # ↓↓↓ 新增：当日收益率，供 7.py 做 IC 检验
    #     必须在 tail() 截断之前算，否则首日收益丢失
    base['收益率'] = df['Close'].pct_change()
    base['收益率'] = base['收益率'].clip(-0.22, 0.22)   # 防复权跳变/退市异常污染均值

    if SIGNAL_RECENT_DAYS:
        base = base.tail(SIGNAL_RECENT_DAYS)

    if not boards:
        boards = [{'板块': '未分类', '类别': '未分类'}]

    frames = []
    for b in boards:
        one = base.copy()
        one.insert(0, '类型', kind)                                  # 板块 / 宽基
        one.insert(1, '归属', region)                                # 国内 / 国外
        one.insert(2, '板块类别', b.get('类别') or '未分类')          # 行业 / 概念 / 宽基
        one.insert(3, '板块', b.get('板块') or '未分类')
        one.insert(4, '股票代码', code)                              # 用于统计板块内股票数量
        frames.append(one)

    return pd.concat(frames) if len(frames) > 1 else frames[0]


def fetch_task(task):
    """【线程池】只做网络抓取，返回原始数据；并发下接口偶发失败，带退避重试"""
    for attempt in range(FETCH_RETRY + 1):
        try:
            df = get_stock_data(task['code'], task.get('name', ''))
            if df is not None and not df.empty:
                return task, df
        except Exception as e:
            if attempt == FETCH_RETRY:
                print(f"❌ {task['name']} ({task['code']}) 抓取异常: {e}")
        if attempt < FETCH_RETRY:
            time.sleep(1.5 * (attempt + 1))   # 退避重试，避免被接口限流
    return task, pd.DataFrame()


def render_task(args):
    """【进程池】形态识别 + 绘图 + 信号汇总，只把很小的信号表传回主进程"""
    task, df, output_dir = args
    try:
        df = extract_pattern_names(df)
        signal_df = build_signal_frame(df, task['kind'], task['region'],
                                       task['boards'], task['code'])
        if GENERATE_CHARTS:
            plot_interactive(df, task['name'], task['code'], task['sector'], output_dir,
                             sector_display=task.get('sector_display', task['sector']))
        return signal_df, None
    except Exception as e:
        return None, f"❌ {task['name']} ({task['code']}) 绘图失败: {e}"


# ----------------------------- 主运行流 -----------------------------
# 注：这个 __main__ 是本模块独立运行时的入口（板块个股全流程），仅用于单独调试。
# 开源仓库里推荐的实际入口是 stocks/stock_analysis_suite.py，它会把本文件当内核模块导入，
# 默认 BROAD_INDEX_ONLY=True 只出宽基指数，不需要 ztjj_board_stocks.csv / target_stocks.csv。
if __name__ == '__main__':
    mp.freeze_support()          # Windows 打包/spawn 必需
    t_start = time.time()

    output_directory = "stock_charts_output"

    if not os.path.exists(output_directory):
        os.makedirs(output_directory)

    # ---------- 1) 读板块池 → 勾选板块（默认概念 + DEFAULT_BOARDS） ----------
    print(">>> 正在读取板块股票池...")
    pool = load_board_pool(BOARD_CSV)
    name2cat = dict(zip(pool['板块'], pool['类别'])) if not pool.empty else {}

    choice = choose_boards(pool)
    boards = choice.get('boards', [])
    top_n = choice.get('top_n', DEFAULT_TOP_N_PER_BOARD)

    if boards:
        shown = '、'.join(boards[:12]) + ('…' if len(boards) > 12 else '')
        print(f">>> 已选【{choice.get('category')}】{len(boards)} 个板块：{shown}")
        print(f">>> 每板块取前 {top_n if top_n else '全部'} 只（按基金持有比例）")
    else:
        print("⚠️ 未选到任何板块，本次只分析国外标的与宽基指数")

    # ---------- 2) 组装标的：国内(ztjj) + 国外(target_stocks，口径不变) ----------
    dom = build_domestic_targets(pool, boards, top_n)
    legacy = build_legacy_targets(LEGACY_CSV, name2cat,
                                  include_overseas=choice.get('overseas', INCLUDE_OVERSEAS),
                                  include_domestic=choice.get('legacy', INCLUDE_LEGACY_DOMESTIC))
    targets = merge_targets(dom, legacy)

    # ---------- 3) 宽基指数照旧压入队列 ----------
    for code, info in BROAD_INDICES.items():
        targets[code] = dict(code=code, name=info[0], region=info[3], kind='宽基',
                             boards=[{'板块': info[4], '类别': '宽基'}])

    # ---------- 4) 整理成任务列表（此时已按代码去重：一只票只抓一次、只画一张图） ----------
    tasks = []
    for t in targets.values():
        bs = t['boards'] or [{'板块': '', '类别': '未分类'}]
        tasks.append(dict(
            code=normalize_code(t['code']),
            name=t['name'],
            sector=bs[0]['板块'],                                  # 主板块：文件名前缀
            sector_display=' / '.join(b['板块'] for b in bs if b['板块']),   # 标题展示全部板块
            boards=bs,                                             # 汇总时按这些板块分发信号
            region=t.get('region', '国内'),
            kind=t.get('kind', '板块'),
        ))

    total = len(tasks)
    multi = sum(1 for t in tasks if len(t['boards']) > 1)
    pair_cnt = sum(len(t['boards']) for t in tasks)
    print(f">>> 共 {total} 个标的（去重后）｜其中 {multi} 只横跨多个板块，"
          f"板块-成分对 {pair_cnt} 条｜抓取并发 {FETCH_WORKERS} 线程，"
          f"绘图并发 {RENDER_WORKERS if USE_PROCESS_POOL else 1} 进程")

    # 用于收集每天每只股票/指数产生的所有信号
    all_signals_list = []
    done = 0

    def collect(signal_df, err):
        """统一处理绘图结果"""
        global done
        done += 1
        if err:
            print(err)
        elif signal_df is not None:
            all_signals_list.append(signal_df)
        if done % 10 == 0 or done == total:
            print(f"    —— 进度 {done}/{total}，已用时 {time.time()-t_start:.1f}s")

    # 抓取(网络IO，线程) 与 绘图(CPU，进程) 组成流水线：
    # 一只票的数据一到就立刻丢给进程池去画，不必等全部抓完
    pool = None
    try:
        if USE_PROCESS_POOL and RENDER_WORKERS > 1:
            pool = ProcessPoolExecutor(max_workers=RENDER_WORKERS)
    except Exception as e:
        print(f"⚠️ 进程池创建失败，退回单进程绘图: {e}")
        pool = None

    render_futures = []
    with ThreadPoolExecutor(max_workers=FETCH_WORKERS) as tp:
        fetch_futures = [tp.submit(fetch_task, t) for t in tasks]

        for fut in as_completed(fetch_futures):
            task, df = fut.result()

            if df is None or df.empty:
                done += 1
                print(f"❌ {task['name']} ({task['code']}) 数据获取失败，跳过。")
                continue

            print(f">>> 已获取: [{task['region']}]【{task.get('sector_display') or task['sector']}】 "
                  f"{task['name']} ({task['code']})")

            if pool is not None:
                render_futures.append(pool.submit(render_task,
                                                  (task, df, output_directory)))
            else:
                sig, err = render_task((task, df, output_directory))
                collect(sig, err)

    if pool is not None:
        for fut in as_completed(render_futures):
            try:
                sig, err = fut.result()
            except Exception as e:
                sig, err = None, f"❌ 绘图子进程异常: {e}"
            collect(sig, err)
        pool.shutdown(wait=True)

    print(f"\n>>> 图表阶段完成，用时 {time.time()-t_start:.1f}s")

    if all_signals_list:
        print("\n>>> 正在聚合全市场信号，生成 Excel 汇总表...")
        combined_df = pd.concat(all_signals_list)
        combined_df.reset_index(inplace=True)

        if 'Date' in combined_df.columns:
            combined_df.rename(columns={'Date': '日期'}, inplace=True)
        elif 'index' in combined_df.columns:
            combined_df.rename(columns={'index': '日期'}, inplace=True)

        combined_df['日期'] = pd.to_datetime(combined_df['日期']).dt.strftime('%Y-%m-%d')

        METRICS = ['看涨', '看跌', '向上缺口', '向下缺口', '量价看多', '量价看空',
                   '涨停', '跌停']

        def build_agg(df_src, keys):
            """按给定维度聚合：数量去重计数 + 各项信号求和 + 重算利好/利空/多空比"""
            cnt = df_src.groupby(keys)['股票代码'].nunique().reset_index(name='板块数量')
            summ = df_src.groupby(keys)[METRICS].sum().reset_index()
            out = pd.merge(summ, cnt, on=keys)

            # ↓↓↓ 新增：板块等权当日收益（成分股收益的算术平均，计数求和、收益取均值）
            if '收益率' in df_src.columns:
                ret = df_src.groupby(keys)['收益率'].mean().reset_index()
                ret.rename(columns={'收益率': '板块收益'}, inplace=True)
                out = pd.merge(out, ret, on=keys, how='left')
            else:
                out['板块收益'] = np.nan

            # 新口径：利好 = 看涨形态 + 向上缺口 + 量价看多
            #         利空 = 看跌形态 + 向下缺口 + 量价看空
            # ---- 口径 v2（2026-08 依据 IC 检验修订）----
            # 看涨/看跌形态未通过稳定性检验（1日 t=-2.35 全由 5 月后单一区制贡献，
            # 10日前后段 IC 反号），留在多空计数里只会稀释量价信号，故剔除。
            # 想回到旧口径把 SIGNAL_SCHEME 改成 'v1' 即可，两种口径可对照跑 IC。
            if SIGNAL_SCHEME == 'v1':
                out['利好'] = out['看涨'] + out['向上缺口'] + out['量价看多']
                out['利空'] = out['看跌'] + out['向下缺口'] + out['量价看空']
            else:
                out['利好'] = out['向上缺口'] + out['量价看多']
                out['利空'] = out['向下缺口'] + out['量价看空']
            out['多空指标比'] = ((out['利好'] - out['利空'])
                                 / out['板块数量'].replace(0, np.nan)).round(4)
            # 涨跌停净额与涨停率：独立于多空口径，任何 SIGNAL_SCHEME 下都一样
            out['涨跌停净'] = out['涨停'] - out['跌停']
            out['涨停率'] = (out['涨停'] / out['板块数量'].replace(0, np.nan)).round(4)

            cols = (keys + ['板块数量'] + METRICS
                    + ['利好', '利空', '多空指标比', '涨跌停净', '涨停率', '板块收益'])
            out = out[cols]
            out.sort_values(by=['日期'] + keys[1:],
                            ascending=[False] + [True] * (len(keys) - 1), inplace=True)
            return out

        if '板块类别' not in combined_df.columns:
            combined_df['板块类别'] = '未分类'

        # 明细表：日期 × 类型 × 归属 × 板块类别 × 板块（最细粒度，HTML 报表以此为数据源）
        detail_df = build_agg(combined_df, ['日期', '类型', '归属', '板块类别', '板块'])

        # 全球合并：同名板块把国内 + 国外合在一起（如 CPO = A股CPO + 美股CPO）
        global_df = build_agg(combined_df, ['日期', '类型', '板块类别', '板块'])
        global_df.insert(2, '归属', '全球合并')

        # 国内 / 国外 单独视图
        dom_df = build_agg(combined_df[combined_df['归属'] == '国内'],
                           ['日期', '类型', '板块类别', '板块'])
        ovs_df = build_agg(combined_df[combined_df['归属'] == '国外'],
                           ['日期', '类型', '板块类别', '板块'])

        # ---- 市场基准：按去重个股统计全市场等权收益与净广度，供 7.py 剔 beta 用 ----
        # drop_duplicates 必不可少：一只票横跨多个板块会被复制多份，不去重会重复计权
        uni = (combined_df[combined_df['类型'] == '板块']
               .drop_duplicates(['日期', '归属', '股票代码']))
        mkt = uni.groupby(['日期', '归属']).agg(
            全市场样本数=('股票代码', 'nunique'),
            全市场收益=('收益率', 'mean'),
            全市场利好=('量价看多', 'sum'),
            全市场利空=('量价看空', 'sum'),
            全市场涨停=('涨停', 'sum'),
            全市场跌停=('跌停', 'sum'),
        ).reset_index()
        mkt['全市场净广度'] = ((mkt['全市场利好'] - mkt['全市场利空'])
                              / mkt['全市场样本数'].replace(0, np.nan)).round(4)
        mkt.sort_values(['日期', '归属'], ascending=[False, True], inplace=True)

        # 口径元信息：6.py 启动时读这张表自动对齐 SIGNAL_SCHEME，
        # 杜绝"表格用 v1、研究用 v2"这类两边各改各的静默错位
        meta = pd.DataFrame([
            {"键": "SIGNAL_SCHEME", "值": SIGNAL_SCHEME},
            {"键": "利好口径", "值": ("看涨+向上缺口+量价看多" if SIGNAL_SCHEME == 'v1'
                                      else "向上缺口+量价看多")},
            {"键": "利空口径", "值": ("看跌+向下缺口+量价看空" if SIGNAL_SCHEME == 'v1'
                                      else "向下缺口+量价看空")},
            {"键": "涨跌停口径", "值": ("收盘封板" if LIMIT_SEAL_ONLY else "含炸板")
                                       + "；主板10%/创业板科创板20%/北交所30%/主板ST5%；"
                                         "境外标的不统计"},
            {"键": "SIGNAL_RECENT_DAYS", "值": SIGNAL_RECENT_DAYS},
            {"键": "生成时间", "值": time.strftime("%Y-%m-%d %H:%M:%S")},
        ])

        excel_path = os.path.join(output_directory, "板块形态统计汇总.xlsx")
        sheets = {
            "元信息": meta,
            "明细": detail_df,          # 供 Sector_heatmap.py 读取
            "全球合并": global_df,
            "国内": dom_df,
            "国外": ovs_df,
            # 按板块类别拆两张，方便直接翻“行业”或“概念”
            "行业": global_df[global_df['板块类别'] == '行业'],
            "概念": global_df[global_df['板块类别'] == '概念'],
            "宽基": detail_df[detail_df['类型'] == '宽基'],
            "市场基准": mkt,                              # 新增：剔 beta 的基准
        }
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            for sname, sdf in sheets.items():
                if sdf is not None and not sdf.empty:
                    sdf.to_excel(writer, sheet_name=sname, index=False)

        try:
            wb = openpyxl.load_workbook(excel_path)

            red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
            green_fill = PatternFill(start_color="99FF99", end_color="99FF99", fill_type="solid")

            for ws in wb.worksheets:
                headers = [cell.value for cell in ws[1]]

                # 利好 / 利空列的基础染色
                if '利好' in headers and '利空' in headers:
                    lihao_idx = headers.index('利好') + 1
                    likong_idx = headers.index('利空') + 1

                    for row in range(2, ws.max_row + 1):
                        lihao_val = ws.cell(row=row, column=lihao_idx).value
                        if isinstance(lihao_val, (int, float)) and lihao_val > 0:
                            ws.cell(row=row, column=lihao_idx).fill = red_fill

                        likong_val = ws.cell(row=row, column=likong_idx).value
                        if isinstance(likong_val, (int, float)) and likong_val > 0:
                            ws.cell(row=row, column=likong_idx).fill = green_fill

                # 多空指标比列条件格式：>0.2 红 / <-0.2 绿
                if '多空指标比' in headers:
                    ratio_col_idx = headers.index('多空指标比') + 1
                    for row in range(2, ws.max_row + 1):
                        val = ws.cell(row=row, column=ratio_col_idx).value
                        if isinstance(val, (int, float)):
                            if val > 0.2:
                                ws.cell(row=row, column=ratio_col_idx).fill = red_fill
                            elif val < -0.2:
                                ws.cell(row=row, column=ratio_col_idx).fill = green_fill

                # 冻结表头 + 列宽自适应，方便直接翻阅
                ws.freeze_panes = "A2"
                for col_i, h in enumerate(headers, start=1):
                    width = max(10, min(16, len(str(h)) * 2 + 4))
                    ws.column_dimensions[ws.cell(row=1, column=col_i).column_letter].width = width

            wb.save(excel_path)
            print(f"✅ 成功生成带条件格式的汇总报表: {excel_path}")
            print(f"   工作表: {' / '.join(sheets.keys())}")
        except Exception as e:
            print(f"✅ 成功生成报表，但单元格染色应用失败: {e}")

    else:
        print("\n❌ 没有数据用于生成汇总报表。[cite: 1]")